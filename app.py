"""
app.py – Dashboard de Reportes de Labor · FYC Calzado
=======================================================
Aplicación Streamlit para visualizar y analizar los reportes de horas
exportados desde Siesa Access en formato Excel (.xlsx).
"""

import streamlit as st
import pandas as pd
import numpy as np
import plotly.express as px
import plotly.graph_objects as go
from pathlib import Path
import base64
import io
import openpyxl

# Banner corporativo con logo
LOGO_PATH = Path(__file__).parent / "COSMOS.jpg.jpeg"
if LOGO_PATH.exists():
    import base64
    logo_b64 = base64.b64encode(LOGO_PATH.read_bytes()).decode()
    logo_html = f"<div style='background:#FFFFFF; padding:6px 8px; border-radius:8px; display:inline-flex; align-items:center; flex-shrink:0; max-width:120px;'><img src='data:image/jpeg;base64,{logo_b64}' style='max-height:85px; max-width:100%; width:auto; height:auto; display:block;'></div>"
else:
    logo_html = "<span style='font-size:1.8rem; font-weight:800; letter-spacing:2px; color:#FFFFFF;'>FYC CALZADO</span>"

st.markdown(
    f"""
    <div style='background:linear-gradient(90deg, #B8927F, #A68070); padding:1.2rem 1.2rem; display:flex;
                flex-wrap:nowrap; align-items:center; gap:1rem; border-bottom:4px solid #C9765F;
                width:100%; box-sizing:border-box; overflow-x:auto; margin-bottom:1.5rem; margin-top:0.8rem;'>
        {logo_html}
        <div style='flex:1; min-width:140px;'>
          <div style='color:#FFFFFF; font-size:0.8rem; letter-spacing:1.2px; opacity:0.9; font-weight:600; white-space:nowrap; overflow:hidden; text-overflow:ellipsis;'>
            SISTEMA DE REPORTES
          </div>
          <div style='color:#FFFFFF; font-size:0.95rem; font-weight:700; opacity:0.85; white-space:nowrap; overflow:hidden; text-overflow:ellipsis;'>
            Informe de Trabajo · Siesa Access
          </div>
        </div>
    </div>
    """,
    unsafe_allow_html=True
)

# Definir ruta del isotipo
ISOTIPO_PATH = Path(__file__).parent / "isotipo-png.png"

with st.sidebar:
    # Zona de marca en el header del sidebar
    import base64 as _b64
    if ISOTIPO_PATH.exists():
        isotipo_b64 = _b64.b64encode(ISOTIPO_PATH.read_bytes()).decode()
        brand_img = f"""<img src='data:image/png;base64,{isotipo_b64}'
            style='max-height:80px; width:auto; display:block; margin:0 auto 12px;
                   filter: drop-shadow(0 3px 10px rgba(0,0,0,0.3)) brightness(1.05);'>"""
    else:
        brand_img = "<div style='font-size:32px; font-weight:900; color:rgba(255,255,255,0.95); letter-spacing:4px; margin-bottom:12px;'>FYC</div>"

    st.markdown(
        f"""
        <div class='sidebar-brand' style='
            background: linear-gradient(150deg, #9E6550 0%, #7A3E2C 60%, #5C2A1C 100%);
            padding: 30px 20px 24px;
            text-align: center;
            border-bottom: 3px solid rgba(0,0,0,0.18);
            position: relative;
            overflow: hidden;
        '>
            <div style='
                position: absolute; top: -20px; right: -20px;
                width: 100px; height: 100px;
                background: rgba(255,255,255,0.04);
                border-radius: 50%;
            '></div>
            <div style='
                position: absolute; bottom: -30px; left: -15px;
                width: 80px; height: 80px;
                background: rgba(255,255,255,0.03);
                border-radius: 50%;
            '></div>
            {brand_img}
            <div class='brand-title' style='
                color: rgba(255,255,255,0.95);
                font-size: 13px;
                font-weight: 800;
                letter-spacing: 3px;
                text-transform: uppercase;
                line-height: 1.2;
            '>FYC Calzado</div>
            <div class='brand-subtitle' style='
                color: rgba(255,255,255,0.55);
                font-size: 9.5px;
                letter-spacing: 1.8px;
                text-transform: uppercase;
                margin-top: 5px;
            '>Sistema de Reportes</div>
        </div>
        """,
        unsafe_allow_html=True,
    )
    
    # SECCIÓN 1: SUBIR ARCHIVO
    st.markdown("### <i class='lucide lucide-upload'></i> Subir archivo", unsafe_allow_html=True)
    uploaded_file = st.file_uploader(
        "Selecciona el archivo Excel exportado desde Siesa",
        type=["xlsx"],
        help="El archivo no se guarda, solo se usa para el análisis actual."
    )

    # Botón para limpiar archivo subido
    if st.button("Limpiar archivo", use_container_width=True, help="Elimina el archivo cargado y reinicia la app", key="btn_clear_file", type="secondary"):
        for k in list(st.session_state.keys()):
            if "file_uploader" in k or "uploaded_file" in k:
                del st.session_state[k]
        st.rerun()

    if uploaded_file is not None:
        excel_path = uploaded_file
        excel_label = f"(archivo subido: {uploaded_file.name})"
    else:
        excel_path = None
        excel_label = "(sin archivo)"

    st.caption(f"Archivo en uso: {excel_label}")

    with st.columns(1)[0]:
        if st.button("Recargar", use_container_width=True, help="Recarga el dashboard con los datos actuales", key="btn_reload", type="primary"):
            st.rerun()

    st.divider()
    
    # SECCIÓN 2: FILTROS PRINCIPALES
    st.markdown("### <i class='lucide lucide-filter'></i> Filtros - Empleados", unsafe_allow_html=True)

# Abreviaciones de columnas de horas con su descripción
HORA_COLS = {
    "JORNADA": "Jornada ordinaria",
    "DO": "D. ordinario",
    "RNO": "Recargo nocturno ord.",
    "HEDO": "H. extra diurna ord.",
    "HENO": "H. extra nocturna ord.",
    "DOM": "Dominical",
    "RNF": "Rec. noct. festivo",
    "HEDF": "H. extra diurna festivo",
    "HENF": "H. extra noct. festivo",
    "FEST": "Festivo",
    "RNDOM": "Rec. noct. dominical",
    "RDOM": "Rec. dominical",
    "ODOM": "Hora extra dominical",
    "ORNF": "Otra rec. noct. festivo",
    "OEDF": "Otra h. ext. diurna fest.",
    "OFEST": "Otras horas festivo",
    "TOTAL": "Total horas",
}

# ─────────────────────────────────────────────
# PARSER
# ─────────────────────────────────────────────


@st.cache_data(show_spinner="Leyendo archivo Siesa…")
def parse_excel(path_or_file) -> pd.DataFrame:
    """
    Lee y transforma el archivo Excel exportado desde Siesa Access.

    El archivo contiene una hoja llamada "ReporteXML" con bloques de filas
    por empleado. Cada bloque comienza con filas que indican "Nombre:",
    "Documento:" y "Grupo:", seguidas de una fila de encabezados de columnas
    y luego filas de datos donde la primera columna es el día de la semana.

    Parámetros
    ----------
    path_or_file : str | Path | UploadedFile
        Ruta al archivo .xlsx o el objeto subido por st.file_uploader.

    Retorna
    -------
    pd.DataFrame
        DataFrame limpio con columnas: Nombre, Documento, Grupo, Fecha,
        Día, Turno, columnas de horas (JORNADA, HEDO, …, TOTAL),
        Semana, Semana_etiqueta, Mes, Mes_num.
        Retorna un DataFrame vacío si no se encontraron registros.
    """
    if hasattr(path_or_file, 'read'):
        wb = openpyxl.load_workbook(path_or_file, data_only=True)
    else:
        wb = openpyxl.load_workbook(str(path_or_file), data_only=True)
    ws = wb["ReporteXML"]
    rows = list(ws.iter_rows(values_only=True))

    records = []
    current_nombre = None
    current_documento = None
    current_grupo = None
    col_map = {}  # colIndex -> colName

    # Nombres de días en español para detectar filas de datos
    DIAS = {"lunes", "martes", "miércoles", "miercoles", "jueves",
            "viernes", "sábado", "sabado", "domingo"}

    for row in rows:
        # Buscar marcador de empleado
        vals = {i: v for i, v in enumerate(row) if v is not None}
        if not vals:
            continue

        # Detectar "Nombre:"
        for v in vals.values():
            sv = str(v).strip()
            if sv.startswith("Nombre:"):
                current_nombre = sv.replace("Nombre:", "").strip()
                col_map = {}
                break

        # Detectar "Documento:"
        for v in vals.values():
            sv = str(v).strip()
            if sv.startswith("Documento:"):
                current_documento = sv.replace("Documento:", "").strip()
                break

        # Detectar "Grupo:"
        for v in vals.values():
            sv = str(v).strip()
            if sv.startswith("Grupo:"):
                current_grupo = sv.replace("Grupo:", "").strip()
                break

        # Detectar fila de encabezados de columnas (contiene 'Fecha' y 'TOTAL')
        row_str_vals = [str(v).strip() for v in vals.values()]
        if "Fecha" in row_str_vals and "TOTAL" in row_str_vals:
            col_map = {}
            for idx, v in vals.items():
                if v is not None:
                    name = str(v).strip().replace("\n", " ")
                    col_map[idx] = name
            continue

        # Detectar fila de datos (primera columna no-nula es un día de la semana)
        if col_map and current_nombre:
            first_val = None
            first_idx = None
            for i in sorted(vals.keys()):
                if vals[i] is not None:
                    first_val = str(vals[i]).strip().lower()
                    first_idx = i
                    break

            if first_val and first_val in DIAS:
                # Es una fila de datos
                rec = {
                    "Nombre": current_nombre,
                    "Documento": current_documento,
                    "Grupo": current_grupo,
                }
                # Mapear columnas según los encabezados dinámicos
                for col_idx, col_name in col_map.items():
                    raw = vals.get(col_idx)
                    if raw is not None and str(raw).strip() not in ("", " "):
                        rec[col_name] = str(raw).strip()
                    else:
                        rec[col_name] = None

                # Normalizar campo Día
                rec["Día"] = str(vals.get(first_idx, "")).strip().capitalize()

                records.append(rec)

    if not records:
        return pd.DataFrame()

    df = pd.DataFrame(records)

    # ── Normalizar fecha ──
    if "Fecha" in df.columns:
        df["Fecha"] = pd.to_datetime(
            df["Fecha"].str.strip(), errors="coerce",
            format="%m/%d/%Y",
        )
        # Intentar formato alternativo si falló
        mask = df["Fecha"].isna()
        if mask.any():
            df.loc[mask, "Fecha"] = pd.to_datetime(
                df.loc[mask, "Fecha_raw"] if "Fecha_raw" in df.columns else df.loc[mask, "Fecha"],
                errors="coerce",
            )

    # ── Semana y mes ──
    if "Fecha" in df.columns:
        df["Semana"] = df["Fecha"].dt.isocalendar().week.astype("Int64")
        df["Semana_etiqueta"] = df["Fecha"].dt.to_period("W").astype(str)
        df["Mes"] = df["Fecha"].dt.strftime("%B %Y")
        df["Mes_num"] = df["Fecha"].dt.month

    # ── Convertir columnas de horas a float ──
    hour_candidates = list(HORA_COLS.keys()) + ["TOTAL"]
    for col in hour_candidates:
        if col in df.columns:
            df[col] = pd.to_numeric(df[col], errors="coerce").fillna(0.0)

    # Asegurar TOTAL: si la columna se llama 'TOTAL' o coincide parcialmente
    total_col = next((c for c in df.columns if c.upper() == "TOTAL"), None)
    if total_col and total_col != "TOTAL":
        df["TOTAL"] = df[total_col]

    # Filtrar filas sin fecha válida
    if "Fecha" in df.columns:
        df = df[df["Fecha"].notna()].copy()

    return df


# ─────────────────────────────────────────────
# CARGA DE DATOS
# ─────────────────────────────────────────────

if uploaded_file is None:
    st.info("📂 Sube un archivo Excel de Siesa Access desde el panel izquierdo para comenzar.")
    st.stop()

try:
    df = parse_excel(excel_path)
except Exception as e:
    st.error(f"No se pudieron extraer datos del archivo. Verifica el formato.\n\nError: {e}")
    st.stop()

if df.empty:
    st.error("No se pudieron extraer datos del archivo. Verifica el formato.")
    st.stop()

# Columnas de horas disponibles en el DataFrame
HORAS_DISPONIBLES = [c for c in HORA_COLS if c in df.columns]
if "TOTAL" in df.columns and "TOTAL" not in HORAS_DISPONIBLES:
    HORAS_DISPONIBLES.append("TOTAL")

# ─────────────────────────────────────────────
# SIDEBAR – FILTROS (continuación)
# ─────────────────────────────────────────────
with st.sidebar:
    # SECCIÓN 2: FILTROS PRINCIPALES
    st.markdown("### 🔎 Filtros - Empleados")
    personas = sorted(df["Nombre"].unique())
    sel_personas = st.multiselect(
        "Empleado(s)",
        personas,
        default=personas,
        help="Selecciona uno o varios empleados",
    )

    st.divider()

    # SECCIÓN 3: PERÍODO DE ANÁLISIS
    st.markdown("### <i class='lucide lucide-calendar'></i> Período y vista", unsafe_allow_html=True)
    
    # Período de análisis
    vista = st.radio(
        "Agrupar por",
        ["Día", "Semana", "Mes"],
        index=0,
    )

    st.divider()

    # SECCIÓN 4: RANGO DE FECHAS
    st.markdown("### <i class='lucide lucide-calendar-range'></i> Rango de fechas", unsafe_allow_html=True)

    st.divider()

    # Rango de fechas
    min_fecha = df["Fecha"].min().date()
    max_fecha = df["Fecha"].max().date()
    rango = st.date_input(
        "Desde — Hasta",
        value=(min_fecha, max_fecha),
        min_value=min_fecha,
        max_value=max_fecha,
    )
    if isinstance(rango, (list, tuple)) and len(rango) == 2:
        fecha_ini, fecha_fin = rango
    else:
        fecha_ini, fecha_fin = min_fecha, max_fecha

    st.divider()

    # SECCIÓN 5: MÉTRICA DE HORAS
    st.markdown("### <i class='lucide lucide-activity'></i> Métrica de horas", unsafe_allow_html=True)
    
    # Tipo de hora
    tipo_hora = st.selectbox(
        "Selecciona métrica",
        HORAS_DISPONIBLES,
        index=HORAS_DISPONIBLES.index("TOTAL") if "TOTAL" in HORAS_DISPONIBLES else 0,
        format_func=lambda x: f"{x} – {HORA_COLS.get(x, '')}",
    )



# ─────────────────────────────────────────────
# APLICAR FILTROS
# ─────────────────────────────────────────────
dff = df[
    df["Nombre"].isin(sel_personas)
    & (df["Fecha"].dt.date >= fecha_ini)
    & (df["Fecha"].dt.date <= fecha_fin)
].copy()

# ─────────────────────────────────────────────
# ENCABEZADO
# ─────────────────────────────────────────────
st.title("FYC Calzado – Reporte de Labor")
st.caption("Fuente: Siesa Access · Octubre 2025")

if dff.empty:
    st.warning("No hay datos para los filtros seleccionados.")
    st.stop()

# Redondear todas las columnas numéricas de horas a máximo 2 decimales
_hora_num_cols = [c for c in list(HORA_COLS.keys()) + ["TOTAL"] if c in dff.columns]
for _c in _hora_num_cols:
    dff[_c] = dff[_c].round(2)

# ─────────────────────────────────────────────
# KPIs
# ─────────────────────────────────────────────
def _col_sum(col): return dff[col].sum() if col in dff.columns else 0

total_horas       = dff["TOTAL"].sum() if "TOTAL" in dff.columns else 0
total_do          = _col_sum("DO")
total_rno         = _col_sum("RNO")
total_hedo        = _col_sum("HEDO")
total_heno        = _col_sum("HENO")
total_dom         = _col_sum("DOM")
total_rnf         = _col_sum("RNF")
total_hedf        = _col_sum("HEDF")
total_henf        = _col_sum("HENF")
total_fest        = _col_sum("FEST")
total_rndom       = _col_sum("RNDOM")
total_rdom        = _col_sum("RDOM")
total_extras_ord  = total_hedo + total_heno
total_extras_fest = total_hedf + total_henf
dias_laborados    = int((dff["DO"] > 0).sum()) if "DO" in dff.columns else len(dff)

# ── Fila 1: resumen principal ──
st.markdown("##### Resumen general")
k1, k2, k3 = st.columns(3)
k1.metric("✅ Total horas", f"{total_horas:,.2f}")
k2.metric("DO – Jornada ordinaria", f"{total_do:,.2f}")
k3.metric("Días laborados", dias_laborados)

# ── Fila 2: horas extras separadas (cada tipo con valor de nómina distinto) ──
st.markdown("##### Horas extras – desglose por tipo de pago")
e1, e2, e3, e4 = st.columns(4)
e1.metric("HEDO – H. Extra diurna ord.",   f"{total_hedo:,.2f}", help="Hora extra diurna en día ordinario")
e2.metric("HENO – H. Extra nocturna ord.", f"{total_heno:,.2f}", help="Hora extra nocturna en día ordinario")
e3.metric("HEDF – H. Extra diurna fest.",  f"{total_hedf:,.2f}", help="Hora extra diurna en día festivo")
e4.metric("HENF – H. Extra noct. fest.",   f"{total_henf:,.2f}", help="Hora extra nocturna en día festivo")

# ── Fila 3: otros tipos de hora ──
st.markdown("##### Otros tipos de hora")
d1, d2, d3, d4, d5 = st.columns(5)
d1.metric("RNO",       f"{total_rno:,.2f}",            help="Recargo nocturno ord.")
d2.metric("DOM",       f"{total_dom:,.2f}",            help="Dominical")
d3.metric("RNF",       f"{total_rnf:,.2f}",            help="Rec. noct. festivo")
d4.metric("FEST",      f"{total_fest:,.2f}",           help="Festivo")
d5.metric("RNDOM/RDOM",f"{(total_rndom + total_rdom):,.2f}", help="Rec. noct. dom. + Rec. dom.")

# ── Verificación de sumatoria ──
# JORNADA es columna de referencia del turno, NO es componente del TOTAL en Siesa.
# Incluirla duplicaría las horas ordinarias (DO ya las cubre).
_cols_verificar = ["DO", "RNO", "HEDO", "HENO", "DOM", "RNF", "HEDF", "HENF",
                   "FEST", "RNDOM", "RDOM", "ODOM", "ORNF", "OEDF", "OFEST"]
_suma_parciales = sum(_col_sum(c) for c in _cols_verificar)
_diff = abs(total_horas - _suma_parciales)
if total_horas > 0 and _diff > 0.5:
    st.warning(
        f"⚠️ Diferencia de {_diff:,.2f} h entre la suma de columnas parciales "
        f"({_suma_parciales:,.2f} h) y el TOTAL del archivo ({total_horas:,.2f} h). "
        f"Verifique si el archivo tiene columnas adicionales no reconocidas."
    )
else:
    st.success(f"✔ Verificación OK: suma parciales = {_suma_parciales:,.2f} h  ═  TOTAL archivo = {total_horas:,.2f} h")

st.divider()

# ── Tarjeta de detalle por empleado (si se filtra uno solo) ──
if len(sel_personas) == 1:
    emp_name = sel_personas[0]
    st.markdown(f"### 👤 Detalle – {emp_name}")

    _detail_cols = [c for c in ["DO", "RNO", "HEDO", "HENO", "DOM", "RNF",
                                 "HEDF", "HENF", "FEST", "RNDOM", "RDOM",
                                 "ODOM", "ORNF", "OEDF", "OFEST", "TOTAL"]
                    if c in dff.columns]

    _emp_totals = {c: dff[c].sum() for c in _detail_cols}

    # Construir tabla de desglose
    _filas = []
    _grupo_map = {
        "DO  – Jornada ordinaria":          ["DO"],
        "RNO – Recargo nocturno ord.":       ["RNO"],
        "HEDO – H. Extra diurna ord.":       ["HEDO"],
        "HENO – H. Extra nocturna ord.":     ["HENO"],
        "DOM  – Dominical":                  ["DOM"],
        "FEST – Festivo":                    ["FEST"],
        "RNF  – Rec. noct. festivo":         ["RNF"],
        "HEDF – H. Extra diurna fest.":      ["HEDF"],
        "HENF – H. Extra noct. fest.":       ["HENF"],
        "RNDOM – Rec. noct. dom.":           ["RNDOM"],
        "RDOM  – Rec. dom.":                 ["RDOM"],
        "Otros recargos":                    ["ODOM", "ORNF", "OEDF", "OFEST"],
    }
    for grupo, cols_g in _grupo_map.items():
        subtotal = sum(_emp_totals.get(c, 0) for c in cols_g)
        if subtotal > 0:
            _filas.append({
                "Grupo": grupo,
                "Columnas": " + ".join(c for c in cols_g if c in dff.columns),
                "Total horas": round(subtotal, 2),
            })
    _total_row = {"Grupo": "🔢 TOTAL GENERAL", "Columnas": "TOTAL",
                  "Total horas": round(_emp_totals.get("TOTAL", 0), 2)}
    _suma_row  = {"Grupo": "∑ Suma parciales", "Columnas": " + ".join(
        c for c in _detail_cols if c != "TOTAL"),
        "Total horas": round(sum(v for k, v in _emp_totals.items() if k != "TOTAL"), 2)}

    _df_detalle = pd.DataFrame(_filas + [_suma_row, _total_row])

    def _color_detalle(row):
        if row["Grupo"] == "🔢 TOTAL GENERAL":
            return ["background-color:#8A2F1F; color:white; font-weight:bold"] * len(row)
        if row["Grupo"] == "∑ Suma parciales":
            return ["background-color:#FFF3F0; color:#3D1A0E; font-weight:bold"] * len(row)
        return [""] * len(row)

    st.dataframe(
        _df_detalle.style
            .apply(_color_detalle, axis=1)
            .format({"Total horas": "{:.2f}"}),
        use_container_width=True,
        hide_index=True,
        height=min(80 + len(_df_detalle) * 40, 500),
    )
    st.divider()

# ─────────────────────────────────────────────
# GRÁFICA PRINCIPAL – según vista seleccionada
# ─────────────────────────────────────────────
st.subheader(f"Horas por empleado · agrupado por {vista.lower()}")

if vista == "Día":
    grp_col = "Fecha"
    grp_label = "Fecha"
    grp_df = dff.groupby(["Nombre", "Fecha"])[tipo_hora].sum().reset_index()
    grp_df["Fecha_str"] = grp_df["Fecha"].dt.strftime("%d/%m")
    fig = px.bar(
        grp_df,
        x="Fecha_str",
        y=tipo_hora,
        color="Nombre",
        barmode="group",
        labels={"Fecha_str": "Día", tipo_hora: "Horas"},
        height=420,
    )
    fig.update_layout(xaxis_title="Día", legend_title="Empleado")

elif vista == "Semana":
    grp_df = dff.groupby(["Nombre", "Semana_etiqueta"])[tipo_hora].sum().reset_index()
    fig = px.bar(
        grp_df,
        x="Semana_etiqueta",
        y=tipo_hora,
        color="Nombre",
        barmode="group",
        labels={"Semana_etiqueta": "Semana", tipo_hora: "Horas"},
        height=420,
    )
    fig.update_layout(xaxis_title="Semana (ISO)", legend_title="Empleado")

else:  # Mes
    grp_df = dff.groupby(["Nombre", "Mes"])[tipo_hora].sum().reset_index()
    fig = px.bar(
        grp_df,
        x="Mes",
        y=tipo_hora,
        color="Nombre",
        barmode="group",
        labels={"Mes": "Mes", tipo_hora: "Horas"},
        height=420,
    )
    fig.update_layout(xaxis_title="Mes", legend_title="Empleado")

fig.update_layout(
    plot_bgcolor="rgba(0,0,0,0)",
    paper_bgcolor="rgba(0,0,0,0)",
    font_size=12,
)
import streamlit.components.v1 as components
from plotly.io import to_html
from streamlit_js_eval import streamlit_js_eval

# --- Detección automática de móvil/tablet usando ancho real del navegador ---
_screen_width = streamlit_js_eval(js_expressions="window.innerWidth", key="_screen_w")
is_mobile = (_screen_width is not None and int(_screen_width) < 700)

# TOP_N empleados en móvil ---
TOP_N = 5

# Definir fig_height ANTES de cualquier uso
fig_height = 420
if is_mobile:
    fig_height = 260

# --- Filtrar top N empleados en móvil ANTES de crear la figura ---
TOP_N = 5
if is_mobile and 'Nombre' in grp_df.columns:
    if 'Semana_etiqueta' in grp_df.columns:
        top_empleados = grp_df.groupby('Nombre')[tipo_hora].sum().nlargest(TOP_N).index
        grp_df = grp_df[grp_df['Nombre'].isin(top_empleados)]
    elif 'Mes' in grp_df.columns:
        top_empleados = grp_df.groupby('Nombre')[tipo_hora].sum().nlargest(TOP_N).index
        grp_df = grp_df[grp_df['Nombre'].isin(top_empleados)]



import numpy as np
# Regenerar la figura con el df filtrado (o el original si no es móvil)
if is_mobile:
    # Gráfica de dona en móvil
    if 'Nombre' in grp_df.columns and tipo_hora in grp_df.columns:
        pie_df = grp_df.groupby('Nombre')[tipo_hora].sum().reset_index()
        pie_df = pie_df[pie_df[tipo_hora] > 0]
        fig = px.pie(
            pie_df,
            values=tipo_hora,
            names="Nombre",
            hole=0.45,
            height=fig_height,
            title=f"Distribución de {tipo_hora} por empleado",
        )
        fig.update_traces(
            textinfo='percent',
            textposition='inside',
            insidetextorientation='radial',
        )
        fig.update_layout(
            legend_title="Empleado",
            font_size=12,
            showlegend=True,
            legend=dict(orientation='v', x=1.02, y=0.5),
        )
    else:
        # Fallback: gráfico vacío
        fig = go.Figure()
        fig.add_annotation(text="Sin datos para mostrar", showarrow=False)
else:
    # Barras verticales en desktop
    if 'Semana_etiqueta' in grp_df.columns:
        fig = px.bar(
            grp_df,
            x="Semana_etiqueta",
            y=tipo_hora,
            color="Nombre",
            barmode="group",
            height=fig_height,
            labels={"Semana_etiqueta": "Semana", tipo_hora: "Horas"},
        )
        fig.update_layout(xaxis_title="Semana (ISO)", legend_title="Empleado")
    elif 'Mes' in grp_df.columns:
        fig = px.bar(
            grp_df,
            x="Mes",
            y=tipo_hora,
            color="Nombre",
            barmode="group",
            height=fig_height,
            labels={"Mes": "Mes", tipo_hora: "Horas"},
        )
        fig.update_layout(xaxis_title="Mes", legend_title="Empleado")
    elif 'Fecha_str' in grp_df.columns:
        fig = px.bar(
            grp_df,
            x="Fecha_str",
            y=tipo_hora,
            color="Nombre",
            barmode="group",
            height=fig_height,
            labels={"Fecha_str": "Día", tipo_hora: "Horas"},
        )
        fig.update_layout(xaxis_title="Día", legend_title="Empleado")

# Ajustes visuales extra para móvil
if is_mobile:
    fig.update_layout(
        font_size=11,
        legend_font_size=10,
        legend_itemwidth=40,
        margin=dict(l=10, r=10, t=40, b=40),
    )
    st.info(f"Mostrando solo los {TOP_N} empleados con más horas para mejor visualización en móvil.")
else:
    fig.update_layout(font_size=12)
fig.update_layout(height=fig_height)
st.plotly_chart(fig, use_container_width=True)

# ─────────────────────────────────────────────
# FUNCIONES DE EXPORTACIÓN
# ─────────────────────────────────────────────
import io
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from fpdf import FPDF

# Columnas de horas extras reconocidas
COLS_HORAS_EXTRA = ["HEDO", "HENO", "HEDF", "HENF"]

# Límites legales (Ley 2466 de 2025 / Código Sustantivo del Trabajo)
LIMITE_DIARIO_HE = 2.0    # máximo 2 horas extras por día
LIMITE_SEMANAL_HE = 12.0  # máximo 12 horas extras por semana


def calcular_cumplimiento(df: pd.DataFrame) -> tuple[pd.DataFrame, pd.DataFrame]:
    """
    Verifica el cumplimiento de los límites legales de horas extras.

    Norma aplicada: Ley 2466 de 2025 / Código Sustantivo del Trabajo.
    - Máximo 2 horas extras diarias por empleado.
    - Máximo 12 horas extras semanales por empleado.

    Parámetros
    ----------
    df : pd.DataFrame
        DataFrame con columnas de horas extras, Nombre, Fecha, Semana_etiqueta.

    Retorna
    -------
    (df_diario, df_semanal) : tuple de DataFrames
        df_diario  → una fila por empleado/día con total extras y estado.
        df_semanal → una fila por empleado/semana con total extras y estado.
        En ambos, la columna "Estado" indica "✅ OK" o "🚨 EXCEDE LÍMITE".
    """
    # Columnas de horas extras que existen en el DataFrame
    extras_cols = [c for c in COLS_HORAS_EXTRA if c in df.columns]

    work = df.copy()
    work["_HE_Total"] = work[extras_cols].sum(axis=1) if extras_cols else 0.0

    # ── Resumen diario ──
    grp_dia = (
        work.groupby(["Nombre", "Fecha"])["_HE_Total"]
        .sum()
        .reset_index()
        .rename(columns={"_HE_Total": "H. Extra del día"})
    )
    grp_dia["Fecha"] = grp_dia["Fecha"].dt.strftime("%d/%m/%Y")
    grp_dia["Límite diario"] = LIMITE_DIARIO_HE
    grp_dia["Exceso diario"] = (grp_dia["H. Extra del día"] - LIMITE_DIARIO_HE).clip(lower=0).round(2)
    grp_dia["Estado"] = grp_dia["H. Extra del día"].apply(
        lambda x: "🚨 EXCEDE LÍMITE" if x > LIMITE_DIARIO_HE else "✅ OK"
    )
    # Excluir días sin horas extras (no aplica verificación)
    grp_dia = grp_dia[grp_dia["H. Extra del día"] > 0].reset_index(drop=True)

    # ── Resumen semanal ──
    if "Semana_etiqueta" in work.columns:
        grp_sem = (
            work.groupby(["Nombre", "Semana_etiqueta"])["_HE_Total"]
            .sum()
            .reset_index()
            .rename(columns={"Semana_etiqueta": "Semana", "_HE_Total": "H. Extra semana"})
        )
    else:
        grp_sem = pd.DataFrame(columns=["Nombre", "Semana", "H. Extra semana"])

    grp_sem["Límite semanal"] = LIMITE_SEMANAL_HE
    grp_sem["Exceso semanal"] = (grp_sem["H. Extra semana"] - LIMITE_SEMANAL_HE).clip(lower=0).round(2)
    grp_sem["Estado"] = grp_sem["H. Extra semana"].apply(
        lambda x: "🚨 EXCEDE LÍMITE" if x > LIMITE_SEMANAL_HE else "✅ OK"
    )
    # Excluir semanas sin horas extras (no aplica verificación)
    grp_sem = grp_sem[grp_sem["H. Extra semana"] > 0].reset_index(drop=True)

    return grp_dia, grp_sem


def excel_cumplimiento(df_diario: pd.DataFrame, df_semanal: pd.DataFrame) -> bytes:
    """
    Genera un Excel con dos hojas: incumplimientos diarios y semanales.
    Las filas con exceso se resaltan en rojo para fácil identificación.

    Retorna
    -------
    bytes
        Contenido binario del .xlsx.
    """
    wb = Workbook()

    COLOR_OK       = "E8F5E9"  # verde claro
    COLOR_EXCEDE   = "FFEBEE"  # rojo claro
    COLOR_HEADER   = "8A2F1F"  # rojo corporativo
    thin = Side(style="thin", color="CCCCCC")
    brd  = Border(left=thin, right=thin, top=thin, bottom=thin)

    def _write_sheet(ws, df, titulo):
        # Fila de título
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(df.columns))
        title_cell = ws.cell(row=1, column=1, value=titulo)
        title_cell.font = Font(bold=True, size=12, color="FFFFFF")
        title_cell.fill = PatternFill("solid", fgColor=COLOR_HEADER)
        title_cell.alignment = Alignment(horizontal="center")

        # Encabezados
        hdr_font = Font(bold=True, color="FFFFFF", size=10)
        hdr_fill = PatternFill("solid", fgColor="5D2010")
        for ci, col in enumerate(df.columns, 1):
            c = ws.cell(row=2, column=ci, value=col)
            c.font = hdr_font
            c.fill = hdr_fill
            c.alignment = Alignment(horizontal="center")
            c.border = brd
        ws.freeze_panes = "A3"

        # Filas de datos
        for ri, (_, row) in enumerate(df.iterrows(), 3):
            es_exceso = "EXCEDE" in str(row.get("Estado", ""))
            fill = PatternFill("solid", fgColor=COLOR_EXCEDE if es_exceso else COLOR_OK)
            for ci, val in enumerate(row, 1):
                c = ws.cell(row=ri, column=ci, value=val)
                c.fill = fill
                c.border = brd
                c.font = Font(size=9, bold=es_exceso, color="B71C1C" if es_exceso else "1B5E20")
                if isinstance(val, float):
                    c.number_format = "0.00"

        # Auto-ancho
        for ci, col in enumerate(df.columns, 1):
            max_len = max(len(str(col)), *(len(str(v)) for v in df.iloc[:, ci - 1]))
            ws.column_dimensions[get_column_letter(ci)].width = min(max_len + 2, 35)

    # Hoja 1 – diario
    ws1 = wb.active
    ws1.title = "Incumplimiento Diario"
    _write_sheet(ws1, df_diario, "Reporte de cumplimiento – Horas extras diarias (máx. 2 h/día)")

    # Hoja 2 – semanal
    ws2 = wb.create_sheet("Incumplimiento Semanal")
    _write_sheet(ws2, df_semanal, "Reporte de cumplimiento – Horas extras semanales (máx. 12 h/semana)")

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()


def df_to_excel_grouped(df, title, group_col="Nombre"):
    """
    Genera un archivo Excel (.xlsx) estructurado y visualmente organizado.

    El archivo tiene:
    - Fila de encabezados con fondo rojo corporativo y texto blanco.
    - Por cada empleado: una fila de encabezado con su nombre (fondo salmon),
      sus registros de datos con bordes, y una fila de TOTAL con la suma
      de columnas numéricas (fondo rosa claro).
    - Fila en blanco entre empleados para facilitar la lectura.
    - Anchos de columna automáticos según el contenido.
    - Primera fila fija (freeze panes) para scroll largo.

    Parámetros
    ----------
    df : pd.DataFrame
        Datos a exportar. Debe contener la columna `group_col`.
    title : str
        Título de la hoja (se trunca a 30 caracteres).
    group_col : str, opcional
        Nombre de la columna de agrupación (default: "Nombre").

    Retorna
    -------
    bytes
        Contenido binario del archivo .xlsx listo para descargar.
    """
    wb = Workbook()
    ws = wb.active
    ws.title = title[:30]

    COLOR_HEADER = "8A2F1F"
    COLOR_GROUP  = "F0DDD6"
    COLOR_TOTAL  = "FFF3F0"

    thin = Side(style="thin", color="CCCCCC")
    brd  = Border(left=thin, right=thin, top=thin, bottom=thin)

    cols = list(df.columns)
    has_group = group_col in cols
    display_cols = [c for c in cols if c != group_col] if has_group else cols

    hdr_fill = PatternFill("solid", fgColor=COLOR_HEADER)
    hdr_font = Font(bold=True, color="FFFFFF", size=10)
    for ci, col in enumerate(display_cols, 1):
        cell = ws.cell(row=1, column=ci, value=col)
        cell.fill = hdr_fill
        cell.font = hdr_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = brd
    ws.row_dimensions[1].height = 18
    ws.freeze_panes = "A2"

    grp_fill = PatternFill("solid", fgColor=COLOR_GROUP)
    grp_font = Font(bold=True, color="3D1A0E", size=10)
    data_font = Font(size=9)
    tot_fill  = PatternFill("solid", fgColor=COLOR_TOTAL)
    tot_font  = Font(bold=True, size=9)

    row_num = 2
    groups = df.groupby(group_col, sort=False) if has_group else [(None, df)]
    for emp, grp in groups:
        if has_group:
            ws.merge_cells(start_row=row_num, start_column=1, end_row=row_num, end_column=len(display_cols))
            cell = ws.cell(row=row_num, column=1, value=str(emp))
            cell.fill = grp_fill
            cell.font = grp_font
            cell.alignment = Alignment(horizontal="left", vertical="center")
            ws.row_dimensions[row_num].height = 16
            row_num += 1

        num_cols_names = [c for c in display_cols if grp[c].dtype.kind in ("f", "i")]

        for _, data_row in grp.iterrows():
            for ci, col in enumerate(display_cols, 1):
                val = data_row[col]
                cell = ws.cell(row=row_num, column=ci, value=val)
                cell.font = data_font
                cell.border = brd
                if isinstance(val, float):
                    cell.number_format = "0.00"
            row_num += 1

        if has_group:
            for ci, col in enumerate(display_cols, 1):
                if col in num_cols_names:
                    val = grp[col].sum()
                    cell = ws.cell(row=row_num, column=ci, value=val)
                    cell.number_format = "0.00"
                else:
                    cell = ws.cell(row=row_num, column=ci, value="TOTAL" if ci == 1 else "")
                cell.fill = tot_fill
                cell.font = tot_font
                cell.border = brd
            row_num += 2

    for ci, col in enumerate(display_cols, 1):
        max_len = len(str(col))
        for row_cells in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=ci, max_col=ci):
            for cell in row_cells:
                if cell.value:
                    max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[get_column_letter(ci)].width = min(max_len + 2, 32)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()

def df_to_pdf(df, title):
    """
    Genera un archivo PDF simple con los datos de un DataFrame en tabla.

    Cada columna tiene ancho proporcional al número de columnas.
    Incluye una fila de título centrado y encabezados de columna con borde.

    Parámetros
    ----------
    df : pd.DataFrame
        Datos a exportar.
    title : str
        Título centrado en la parte superior del PDF.

    Retorna
    -------
    bytes
        Contenido binario del archivo .pdf listo para descargar.
    """
    pdf = FPDF()
    pdf.add_page()
    pdf.set_font("Arial", "B", 12)
    pdf.cell(0, 10, title, ln=1, align="C")
    pdf.set_font("Arial", size=8)
    col_width = pdf.w / (len(df.columns) + 1)
    row_height = pdf.font_size * 1.5
    for col in df.columns:
        pdf.cell(col_width, row_height, str(col), border=1)
    pdf.ln(row_height)
    for _, row in df.iterrows():
        for val in row:
            pdf.cell(col_width, row_height, str(val), border=1)
        pdf.ln(row_height)
    out = pdf.output(dest="S")
    if isinstance(out, str):
        return out.encode("latin1")
    if isinstance(out, bytearray):
        return bytes(out)
    return out

# ─────────────────────────────────────────────
# TABLAS RESUMEN
# ─────────────────────────────────────────────
tabs = st.tabs(["Por día", "Por semana", "Por mes", "Detalle completo", "⚖️ Cumplimiento Ley 2466", "Total laborado"])

# ── Tab: Por día ──
with tabs[0]:

    st.subheader("Resumen por persona / día")
    cols_show = ["Nombre", "Fecha", "Día", "Turno"] + [
        c for c in ["DO", "RNO", "HEDO", "HENO", "DOM", "RNF",
                    "HEDF", "HENF", "FEST", "RNDOM", "RDOM",
                    "ODOM", "ORNF", "OEDF", "OFEST", "TOTAL"]
        if c in dff.columns
    ]
    t_dia = dff[cols_show].copy()
    t_dia["Fecha"] = t_dia["Fecha"].dt.strftime("%d/%m/%Y")

    # Filtro extra por persona dentro de la tab
    per_dia = st.selectbox("Filtrar empleado (día)", ["Todos"] + sorted(dff["Nombre"].unique()), key="sel_dia")
    if per_dia != "Todos":
        t_dia = t_dia[t_dia["Nombre"] == per_dia]

    t_dia_sorted = t_dia.sort_values(["Nombre", "Fecha"]).reset_index(drop=True)
    _num_dia = [c for c in t_dia_sorted.columns if t_dia_sorted[c].dtype.kind in ("f", "i") and c != "Fecha"]
    _fmt_dia = {c: "{:.2f}" for c in _num_dia}
    # Un solo empleado: ocultar columna Nombre y mostrar como título
    if per_dia != "Todos":
        st.markdown(f"#### 👤 {per_dia}")
        st.dataframe(t_dia_sorted.drop(columns=["Nombre"]).style.format(_fmt_dia), use_container_width=True, height=400)
    else:
        st.dataframe(t_dia_sorted.style.format(_fmt_dia), use_container_width=True, height=400)

    # Exportar Excel estructurado
    xlsx_dia = df_to_excel_grouped(t_dia_sorted, "Por día")
    st.download_button(
        label="⬇ Descargar Excel",
        data=xlsx_dia,
        file_name="reporte_labor_dia.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )

# ── Tab: Por semana ──
with tabs[1]:

    st.subheader("Resumen por persona / semana")
    hora_cols_sum = [c for c in ["JORNADA", "DO", "RNO", "HEDO", "HENO", "DOM",
                                  "RNF", "HEDF", "HENF", "FEST", "RNDOM", "RDOM", "OFEST", "TOTAL"]
                     if c in dff.columns]
    t_sem = (
        dff.groupby(["Nombre", "Semana_etiqueta"])[hora_cols_sum]
        .sum()
        .reset_index()
        .rename(columns={"Semana_etiqueta": "Semana"})
    )
    per_sem = st.selectbox("Filtrar empleado (semana)", ["Todos"] + sorted(dff["Nombre"].unique()), key="sel_sem")
    if per_sem != "Todos":
        t_sem = t_sem[t_sem["Nombre"] == per_sem]

    t_sem_sorted = t_sem.sort_values(["Nombre", "Semana"]).reset_index(drop=True)
    num_cols = [c for c in t_sem_sorted.columns if c not in ("Nombre", "Semana")]
    if per_sem != "Todos":
        st.markdown(f"#### 👤 {per_sem}")
        df_sem_disp = t_sem_sorted.drop(columns=["Nombre"])
        st.dataframe(df_sem_disp.style.format({c: "{:.2f}" for c in num_cols}), use_container_width=True, height=400)
    else:
        st.dataframe(t_sem_sorted.style.format({c: "{:.2f}" for c in num_cols}), use_container_width=True, height=400)

    # Exportar Excel estructurado
    xlsx_sem = df_to_excel_grouped(t_sem_sorted, "Por semana")
    st.download_button(
        label="⬇ Descargar Excel",
        data=xlsx_sem,
        file_name="reporte_labor_semana.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )

# ── Tab: Por mes ──
with tabs[2]:

    st.subheader("Resumen por persona / mes")
    t_mes = (
        dff.groupby(["Nombre", "Mes"])[hora_cols_sum]
        .sum()
        .reset_index()
    )
    per_mes = st.selectbox("Filtrar empleado (mes)", ["Todos"] + sorted(dff["Nombre"].unique()), key="sel_mes")
    if per_mes != "Todos":
        t_mes = t_mes[t_mes["Nombre"] == per_mes]

    t_mes_sorted = t_mes.sort_values("Nombre").reset_index(drop=True)
    num_cols_mes = [c for c in t_mes_sorted.columns if c not in ("Nombre", "Mes")]
    if per_mes != "Todos":
        st.markdown(f"#### 👤 {per_mes}")
        df_mes_disp = t_mes_sorted.drop(columns=["Nombre"])
        st.dataframe(df_mes_disp.style.format({c: "{:.2f}" for c in num_cols_mes}), use_container_width=True, height=400)
    else:
        st.dataframe(t_mes_sorted.style.format({c: "{:.2f}" for c in num_cols_mes}), use_container_width=True, height=400)

    # Exportar Excel estructurado
    xlsx_mes = df_to_excel_grouped(t_mes_sorted, "Por mes")
    st.download_button(
        label="⬇ Descargar Excel",
        data=xlsx_mes,
        file_name="reporte_labor_mes.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )

# ── Tab: Detalle completo ──
with tabs[3]:

    st.subheader("Datos completos")
    det = dff.copy()
    det["Fecha"] = det["Fecha"].dt.strftime("%d/%m/%Y")
    per_det = st.selectbox("Filtrar empleado (detalle)", ["Todos"] + sorted(dff["Nombre"].unique()), key="sel_det")
    if per_det != "Todos":
        det = det[det["Nombre"] == per_det]

    det_sorted = det.sort_values(["Nombre", "Fecha"]).reset_index(drop=True)
    _num_det = [c for c in det_sorted.columns if det_sorted[c].dtype.kind in ("f", "i") and c != "Fecha"]
    _fmt_det = {c: "{:.2f}" for c in _num_det}
    if per_det != "Todos":
        st.markdown(f"#### 👤 {per_det}")
        st.dataframe(det_sorted.drop(columns=["Nombre"]).style.format(_fmt_det), use_container_width=True, height=500)
    else:
        st.dataframe(det_sorted.style.format(_fmt_det), use_container_width=True, height=500)

    # Botón de descarga Excel
    xlsx_det = df_to_excel_grouped(det_sorted, "Detalle completo")
    st.download_button(
        label="⬇ Descargar Excel",
        data=xlsx_det,
        file_name="reporte_labor_detalle.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )

# ── Tab: Cumplimiento Ley 2466 ──
with tabs[4]:
    st.subheader("⚖️ Cumplimiento normativo – Horas extras")
    st.markdown(
        """
        **Norma aplicada:** Ley 2466 de 2025 / Código Sustantivo del Trabajo
        - 🔴 Máximo **2 horas extras** por día por empleado.
        - 🔴 Máximo **12 horas extras** por semana por empleado.

        Las filas resaltadas en 🚨 indican incumplimiento del límite legal.
        """,
    )

    cump_diario, cump_semanal = calcular_cumplimiento(dff)
    # Redondear columnas numéricas del reporte de cumplimiento
    for _col in ["H. Extra del día", "Exceso diario"]:
        if _col in cump_diario.columns:
            cump_diario[_col] = cump_diario[_col].round(2)
    for _col in ["H. Extra semana", "Exceso semanal"]:
        if _col in cump_semanal.columns:
            cump_semanal[_col] = cump_semanal[_col].round(2)

    # ── Filtro por empleado ──
    per_cump = st.selectbox(
        "Filtrar empleado",
        ["Todos"] + sorted(dff["Nombre"].unique()),
        key="sel_cump",
    )
    if per_cump != "Todos":
        cump_diario  = cump_diario[cump_diario["Nombre"] == per_cump]
        cump_semanal = cump_semanal[cump_semanal["Nombre"] == per_cump]

    # ── Resumen de alertas ──
    n_excesos_dia = (cump_diario["Estado"] == "🚨 EXCEDE LÍMITE").sum()
    n_excesos_sem = (cump_semanal["Estado"] == "🚨 EXCEDE LÍMITE").sum()
    c1, c2 = st.columns(2)
    with c1:
        if n_excesos_dia > 0:
            st.error(f"🚨 {n_excesos_dia} día(s) con más de 2 h extras")
        else:
            st.success("✅ Sin infracciones diarias")
    with c2:
        if n_excesos_sem > 0:
            st.error(f"🚨 {n_excesos_sem} semana(s) con más de 12 h extras")
        else:
            st.success("✅ Sin infracciones semanales")

    # ── Filtro de vista: solo incumplimientos o todos ──
    solo_incumplimientos = st.toggle(
        "Mostrar solo incumplimientos",
        value=True,
        help="Activa para ver únicamente los registros que exceden el límite legal.",
    )

    # Colores condicionales: rojo si excede, verde si OK
    def _color_estado(val):
        if "EXCEDE" in str(val):
            return "background-color: #FFEBEE; color: #B71C1C; font-weight: bold"
        return "background-color: #E8F5E9; color: #1B5E20"

    # ── Tabla diaria ──
    st.markdown("#### Detalle diario")
    disp_dia = cump_diario.copy()
    if solo_incumplimientos:
        disp_dia = disp_dia[disp_dia["Estado"] == "🚨 EXCEDE LÍMITE"]
    if per_cump != "Todos":
        disp_dia = disp_dia.drop(columns=["Nombre"])
    if disp_dia.empty:
        st.success("✅ Sin infracciones diarias en el período seleccionado.")
    else:
        _num_dia_c = [c for c in disp_dia.columns if disp_dia[c].dtype.kind in ("f", "i")]
        st.dataframe(
            disp_dia.style
                .format({c: "{:.2f}" for c in _num_dia_c})
                .applymap(_color_estado, subset=["Estado"]),
            use_container_width=True,
            height=min(60 + len(disp_dia) * 36, 420),
        )

    # ── Tabla semanal ──
    st.markdown("#### Detalle semanal")
    disp_sem = cump_semanal.copy()
    if solo_incumplimientos:
        disp_sem = disp_sem[disp_sem["Estado"] == "🚨 EXCEDE LÍMITE"]
    if per_cump != "Todos":
        disp_sem = disp_sem.drop(columns=["Nombre"])
    if disp_sem.empty:
        st.success("✅ Sin infracciones semanales en el período seleccionado.")
    else:
        _num_sem_c = [c for c in disp_sem.columns if disp_sem[c].dtype.kind in ("f", "i")]
        st.dataframe(
            disp_sem.style
                .format({c: "{:.2f}" for c in _num_sem_c})
                .applymap(_color_estado, subset=["Estado"]),
            use_container_width=True,
            height=min(60 + len(disp_sem) * 36, 360),
        )

    # ── Descarga Excel con las dos hojas (siempre exporta todo, no filtrado) ──
    xlsx_cump = excel_cumplimiento(cump_diario, cump_semanal)
    st.download_button(
        label="⬇ Descargar reporte de cumplimiento (Excel)",
        data=xlsx_cump,
        file_name="cumplimiento_horas_extras.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )

# ── Tab: Total laborado ──
with tabs[5]:
    st.subheader("Total de tiempo laborado")

    # JORNADA se excluye porque es columna de referencia del turno, no componente del TOTAL.
    # El TOTAL de Siesa = DO + recargos + extras + dom + festivos (sin JORNADA).
    _all_hour_cols_t = [c for c in HORA_COLS.keys() if c not in ("TOTAL", "JORNADA") and c in dff.columns]
    _extras_ord_cols  = [c for c in ["HEDO", "HENO"] if c in dff.columns]
    _extras_fest_cols = [c for c in ["HEDF", "HENF"] if c in dff.columns]
    _dom_fest_cols    = [c for c in ["DOM", "FEST", "RNDOM", "RDOM"] if c in dff.columns]
    _recargo_cols     = [c for c in ["RNO", "RNF", "ODOM", "ORNF", "OEDF", "OFEST"] if c in dff.columns]

    _grp_t = dff.groupby("Nombre")[_all_hour_cols_t + (["TOTAL"] if "TOTAL" in dff.columns else [])].sum().reset_index()

    _grp_t["DO"]               = _grp_t["DO"].round(2) if "DO" in _grp_t.columns else 0
    _grp_t["Extras_Ordinarias"] = _grp_t[[c for c in _extras_ord_cols if c in _grp_t.columns]].sum(axis=1).round(2)
    _grp_t["Extras_Festivas"]   = _grp_t[[c for c in _extras_fest_cols if c in _grp_t.columns]].sum(axis=1).round(2)
    _grp_t["Dom_Festivo"]       = _grp_t[[c for c in _dom_fest_cols if c in _grp_t.columns]].sum(axis=1).round(2)
    _grp_t["Recargos"]          = _grp_t[[c for c in _recargo_cols if c in _grp_t.columns]].sum(axis=1).round(2)
    if "TOTAL" in _grp_t.columns:
        _grp_t["TOTAL"] = _grp_t["TOTAL"].round(2)

    _fmt_t = {c: "{:.2f}" for c in _grp_t.columns if _grp_t[c].dtype.kind in ("f", "i")}

    subt = st.tabs(["Resumen", "Ordinario vs Extra", "Detalle completo"])

    # ── Subtab: Resumen agrupado ──
    with subt[0]:
        # Columnas individuales — incluye también ODOM/ORNF/OEDF/OFEST (algunos archivos las tienen)
        _he_cols_exist = [c for c in ["DO", "RNO", "HEDO", "HENO", "DOM", "FEST",
                                      "RNF", "HEDF", "HENF", "RNDOM", "RDOM",
                                      "ODOM", "ORNF", "OEDF", "OFEST"] if c in _grp_t.columns]
        _summary_cols = ["Nombre"] + _he_cols_exist
        if "TOTAL" in _grp_t.columns:
            _summary_cols.append("TOTAL")
        _disp_res = _grp_t[_summary_cols].sort_values("Nombre").reset_index(drop=True)
        _disp_res = _disp_res.rename(columns={
            "DO":    "DO (Jornada ord.)",
            "RNO":   "RNO (Rec. noct. ord.)",
            "HEDO":  "HEDO (H.E. Diurna Ord.)",
            "HENO":  "HENO (H.E. Noct. Ord.)",
            "DOM":   "DOM (Dominical)",
            "FEST":  "FEST (Festivo)",
            "RNF":   "RNF (Rec. noct. fest.)",
            "HEDF":  "HEDF (H.E. Diurna Fest.)",
            "HENF":  "HENF (H.E. Noct. Fest.)",
            "RNDOM": "RNDOM (Rec. noct. dom.)",
            "RDOM":  "RDOM (Rec. dom.)",
            "ODOM":  "ODOM (H.E. dom.)",
            "ORNF":  "ORNF (Rec. noct. fest.)",
            "OEDF":  "OEDF (H.E. diurna fest.)",
            "OFEST": "OFEST (Otras fest.)",
            "TOTAL": "TOTAL",
        })
        st.caption("Cada columna es la suma de todas las horas del período. TOTAL = suma de todas las categorías.")
        st.dataframe(_disp_res.style.format({c: "{:.2f}" for c in _disp_res.columns if _disp_res[c].dtype.kind in ("f","i")}),
                     use_container_width=True, height=min(60 + len(_disp_res) * 36, 500))
        _xlsx_res = df_to_excel_grouped(_disp_res, "Resumen laborado")
        st.download_button("⬇ Descargar Excel", data=_xlsx_res, file_name="total_resumen.xlsx",
                           mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", key="dl_tot_res")

    # ── Subtab: Ordinario vs Extra ──
    with subt[1]:
        _norm_cols = [c for c in _all_hour_cols_t if c not in (_extras_ord_cols + _extras_fest_cols)]
        _grp_t["Tiempo_Ordinario"] = _grp_t[[c for c in _norm_cols if c in _grp_t.columns]].sum(axis=1).round(2)
        _grp_t["Tiempo_Extra"]     = (_grp_t["Extras_Ordinarias"] + _grp_t["Extras_Festivas"]).round(2)
        _disp_ord = _grp_t[["Nombre", "Tiempo_Ordinario", "Tiempo_Extra"] + (["TOTAL"] if "TOTAL" in _grp_t.columns else [])].sort_values("Nombre").reset_index(drop=True)
        st.dataframe(_disp_ord.style.format({c: "{:.2f}" for c in _disp_ord.columns if _disp_ord[c].dtype.kind in ("f","i")}),
                     use_container_width=True, height=min(60 + len(_disp_ord) * 36, 500))
        _xlsx_ord = df_to_excel_grouped(_disp_ord, "Total Ordinario vs Extra")
        st.download_button("⬇ Descargar Excel", data=_xlsx_ord, file_name="total_ordinario_extra.xlsx",
                           mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", key="dl_tot_ord")

    # ── Subtab: Detalle completo ──
    with subt[2]:
        _det_cols = ["Nombre"] + [c for c in _all_hour_cols_t if c in _grp_t.columns] + (["TOTAL"] if "TOTAL" in _grp_t.columns else [])
        _disp_ext = _grp_t[_det_cols].sort_values("Nombre").reset_index(drop=True)
        st.caption("Todas las columnas de horas individuales del reporte Siesa.")
        st.dataframe(_disp_ext.style.format({c: "{:.2f}" for c in _disp_ext.columns if _disp_ext[c].dtype.kind in ("f","i")}),
                     use_container_width=True, height=min(60 + len(_disp_ext) * 36, 500))
        _xlsx_ext = df_to_excel_grouped(_disp_ext, "Detalle completo laborado")
        st.download_button("⬇ Descargar Excel", data=_xlsx_ext, file_name="total_detalle_completo.xlsx",
                           mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", key="dl_tot_ext")

# ─────────────────────────────────────────────
# GRÁFICA COMPARATIVA – Distribución de tipos de horas
# ─────────────────────────────────────────────
st.divider()
st.subheader("Distribución de tipos de horas por empleado")

dist_cols = [c for c in ["JORNADA", "HEDO", "HENO", "DOM", "HEDF", "HENF", "FEST"] if c in dff.columns]

# --- Definir altura de la gráfica antes de usarla ---
fig2_height = 420
if is_mobile:
    fig2_height = 260

dist_df = dff.groupby("Nombre")[dist_cols].sum().reset_index()
dist_melt = dist_df.melt(id_vars="Nombre", var_name="Tipo", value_name="Horas")
dist_melt["Tipo_desc"] = dist_melt["Tipo"].map(lambda x: HORA_COLS.get(x, x))
dist_melt_filtered = dist_melt[dist_melt["Horas"] > 0]

# --- Lógica para la segunda gráfica ---
if is_mobile:
    # Mostrar solo top N empleados en móvil, siempre como dona/pie
    if dist_melt_filtered.shape[0] > 0 and 'Nombre' in dist_melt_filtered.columns:
        top_empleados2 = dist_melt_filtered.groupby('Nombre')['Horas'].sum().nlargest(TOP_N).index
        dist_melt_filtered_mobile = dist_melt_filtered[dist_melt_filtered['Nombre'].isin(top_empleados2)]
        fig2 = px.pie(
            dist_melt_filtered_mobile,
            values="Horas",
            names="Nombre",
            hole=0.45,
            height=fig2_height,
            title=f"Distribución de horas por empleado (top {TOP_N})",
        )
        fig2.update_traces(
            textinfo='percent',
            textposition='inside',
            insidetextorientation='radial',
        )
        fig2.update_layout(
            legend_title="Empleado",
            font_size=12,
            showlegend=True,
            legend=dict(orientation='v', x=1.02, y=0.5),
        )
        st.info(f"Mostrando solo los {TOP_N} empleados con más horas para mejor visualización en móvil.")
    else:
        fig2 = px.pie()  # Gráfica vacía si no hay datos
else:
    # Desktop: dona si solo hay un empleado, barras apiladas si hay varios
    if dist_df.shape[0] == 1:
        fig2 = px.pie(
            dist_melt_filtered,
            values="Horas",
            names="Tipo_desc",
            hole=0.4,
            labels={"Horas": "Horas", "Tipo_desc": "Tipo"},
            height=fig2_height,
            title="Composición de horas por empleado (período seleccionado)",
        )
        fig2.update_layout(
            plot_bgcolor="rgba(0,0,0,0)",
            paper_bgcolor="rgba(0,0,0,0)",
            legend_title="Tipo de hora",
        )
    else:
        fig2 = px.bar(
            dist_melt_filtered,
            x="Nombre",
            y="Horas",
            color="Tipo_desc",
            barmode="stack",
            height=fig2_height,
            labels={"Nombre": "Empleado", "Horas": "Horas", "Tipo_desc": "Tipo"},
            title="Composición de horas por empleado (período seleccionado)",
        )
        fig2.update_layout(
            xaxis_tickangle=-30,
            plot_bgcolor="rgba(0,0,0,0)",
            paper_bgcolor="rgba(0,0,0,0)",
            legend_title="Tipo de hora",
            font_size=12,
        )

fig2.update_layout(height=fig2_height)
st.plotly_chart(fig2, use_container_width=True)

st.caption("Desarrollado con Streamlit · Datos: Siesa Access")
