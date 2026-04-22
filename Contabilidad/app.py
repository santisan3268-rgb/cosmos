"""
app.py – Dashboard de Reportes de Labor · FYC Calzado
=======================================================
Aplicación Streamlit para visualizar y analizar los reportes de horas
exportados desde Siesa Access en formato Excel (.xlsx).
"""

import streamlit as st
import pandas as pd
import numpy as np
import plotly.express as px
import plotly.graph_objects as go
from pathlib import Path
import base64
import io
import openpyxl
import os

try:
    import pytds
except ImportError:
    pytds = None

# Banner corporativo con logo
LOGO_PATH = Path(__file__).parent / "COSMOS.jpg.jpeg"
if LOGO_PATH.exists():
    import base64
    logo_b64 = base64.b64encode(LOGO_PATH.read_bytes()).decode()
    logo_html = f"<div style='background:#FFFFFF; padding:6px 8px; border-radius:8px; display:inline-flex; align-items:center; flex-shrink:0; max-width:120px;'><img src='data:image/jpeg;base64,{logo_b64}' style='max-height:85px; max-width:100%; width:auto; height:auto; display:block;'></div>"
else:
    logo_html = "<span style='font-size:1.8rem; font-weight:800; letter-spacing:2px; color:#FFFFFF;'>FYC CALZADO</span>"

st.markdown(
    f"""
    <div style='background:linear-gradient(90deg, #B8927F, #A68070); padding:1.2rem 1.2rem; display:flex;
                flex-wrap:nowrap; align-items:center; gap:1rem; border-bottom:4px solid #C9765F;
                width:100%; box-sizing:border-box; overflow-x:auto; margin-bottom:1.5rem; margin-top:0.8rem;'>
        {logo_html}
        <div style='flex:1; min-width:140px;'>
          <div style='color:#FFFFFF; font-size:0.8rem; letter-spacing:1.2px; opacity:0.9; font-weight:600; white-space:nowrap; overflow:hidden; text-overflow:ellipsis;'>
            SISTEMA DE REPORTES
          </div>
          <div style='color:#FFFFFF; font-size:0.95rem; font-weight:700; opacity:0.85; white-space:nowrap; overflow:hidden; text-overflow:ellipsis;'>
            Informe de Trabajo · Siesa Access
          </div>
        </div>
    </div>
    """,
    unsafe_allow_html=True
)

# Definir ruta del isotipo
ISOTIPO_PATH = Path(__file__).parent / "isotipo-png.png"

with st.sidebar:
    # Zona de marca en el header del sidebar
    import base64 as _b64
    if ISOTIPO_PATH.exists():
        isotipo_b64 = _b64.b64encode(ISOTIPO_PATH.read_bytes()).decode()
        brand_img = f"""<img src='data:image/png;base64,{isotipo_b64}'
            style='max-height:80px; width:auto; display:block; margin:0 auto 12px;
                   filter: drop-shadow(0 3px 10px rgba(0,0,0,0.3)) brightness(1.05);'>"""
    else:
        brand_img = "<div style='font-size:32px; font-weight:900; color:rgba(255,255,255,0.95); letter-spacing:4px; margin-bottom:12px;'>FYC</div>"

    st.markdown(
        f"""
        <div class='sidebar-brand' style='
            background: linear-gradient(150deg, #9E6550 0%, #7A3E2C 60%, #5C2A1C 100%);
            padding: 30px 20px 24px;
            text-align: center;
            border-bottom: 3px solid rgba(0,0,0,0.18);
            position: relative;
            overflow: hidden;
        '>
            <div style='
                position: absolute; top: -20px; right: -20px;
                width: 100px; height: 100px;
                background: rgba(255,255,255,0.04);
                border-radius: 50%;
            '></div>
            <div style='
                position: absolute; bottom: -30px; left: -15px;
                width: 80px; height: 80px;
                background: rgba(255,255,255,0.03);
                border-radius: 50%;
            '></div>
            {brand_img}
            <div class='brand-title' style='
                color: rgba(255,255,255,0.95);
                font-size: 13px;
                font-weight: 800;
                letter-spacing: 3px;
                text-transform: uppercase;
                line-height: 1.2;
            '>FYC Calzado</div>
            <div class='brand-subtitle' style='
                color: rgba(255,255,255,0.55);
                font-size: 9.5px;
                letter-spacing: 1.8px;
                text-transform: uppercase;
                margin-top: 5px;
            '>Sistema de Reportes</div>
        </div>
        """,
        unsafe_allow_html=True,
    )
    
    # SECCIÓN 1: SUBIR ARCHIVO
    st.markdown("### <i class='lucide lucide-upload'></i> Subir archivo", unsafe_allow_html=True)
    uploaded_file = st.file_uploader(
        "Selecciona el archivo Excel exportado desde Siesa",
        type=["xlsx"],
        help="El archivo no se guarda, solo se usa para el análisis actual."
    )

    # Botón para limpiar archivo subido
    if st.button("Limpiar archivo", use_container_width=True, help="Elimina el archivo cargado y reinicia la app", key="btn_clear_file", type="secondary"):
        for k in list(st.session_state.keys()):
            if "file_uploader" in k or "uploaded_file" in k:
                del st.session_state[k]
        st.rerun()

    if uploaded_file is not None:
        excel_path = uploaded_file
        excel_label = f"(archivo subido: {uploaded_file.name})"
    else:
        excel_path = None
        excel_label = "(sin archivo)"

    st.caption(f"Archivo en uso: {excel_label}")

    with st.columns(1)[0]:
        if st.button("Recargar", use_container_width=True, help="Recarga el dashboard con los datos actuales", key="btn_reload", type="primary"):
            st.rerun()

    st.divider()
    
    # SECCIÓN 2: FILTROS PRINCIPALES
    st.markdown("### <i class='lucide lucide-filter'></i> Filtros - Empleados", unsafe_allow_html=True)

# Abreviaciones de columnas de horas con su descripción
HORA_COLS = {
    "JORNADA": "Jornada ordinaria",
    "DO": "D. ordinario",
    "RNO": "Recargo nocturno ord.",
    "HEDO": "H. extra diurna ord.",
    "HENO": "H. extra nocturna ord.",
    "DOM": "Dominical",
    "RNF": "Rec. noct. festivo",
    "HEDF": "H. extra diurna festivo",
    "HENF": "H. extra noct. festivo",
    "FEST": "Festivo",
    "RNDOM": "Rec. noct. dominical",
    "RDOM": "Rec. dominical",
    "ODOM": "Hora extra dominical",
    "ORNF": "Otra rec. noct. festivo",
    "OEDF": "Otra h. ext. diurna fest.",
    "OFEST": "Otras horas festivo",
    "TOTAL": "Total horas",
}

# ─────────────────────────────────────────────
# PARSER
# ─────────────────────────────────────────────


@st.cache_data(show_spinner="Leyendo archivo Siesa…")
def parse_excel(path_or_file) -> pd.DataFrame:
    """
    Lee y transforma el archivo Excel exportado desde Siesa Access.

    El archivo contiene una hoja llamada "ReporteXML" con bloques de filas
    por empleado. Cada bloque comienza con filas que indican "Nombre:",
    "Documento:" y "Grupo:", seguidas de una fila de encabezados de columnas
    y luego filas de datos donde la primera columna es el día de la semana.

    Parámetros
    ----------
    path_or_file : str | Path | UploadedFile
        Ruta al archivo .xlsx o el objeto subido por st.file_uploader.

    Retorna
    -------
    pd.DataFrame
        DataFrame limpio con columnas: Nombre, Documento, Grupo, Fecha,
        Día, Turno, columnas de horas (JORNADA, HEDO, …, TOTAL),
        Semana, Semana_etiqueta, Mes, Mes_num.
        Retorna un DataFrame vacío si no se encontraron registros.
    """
    if hasattr(path_or_file, 'read'):
        wb = openpyxl.load_workbook(path_or_file, data_only=True)
    else:
        wb = openpyxl.load_workbook(str(path_or_file), data_only=True)
    ws = wb["ReporteXML"]
    rows = list(ws.iter_rows(values_only=True))

    records = []
    current_nombre = None
    current_documento = None
    current_grupo = None
    col_map = {}  # colIndex -> colName

    # Nombres de días en español para detectar filas de datos
    DIAS = {"lunes", "martes", "miércoles", "miercoles", "jueves",
            "viernes", "sábado", "sabado", "domingo"}

    for row in rows:
        # Buscar marcador de empleado
        vals = {i: v for i, v in enumerate(row) if v is not None}
        if not vals:
            continue

        # Detectar "Nombre:"
        for v in vals.values():
            sv = str(v).strip()
            if sv.startswith("Nombre:"):
                current_nombre = sv.replace("Nombre:", "").strip()
                col_map = {}
                break

        # Detectar "Documento:"
        for v in vals.values():
            sv = str(v).strip()
            if sv.startswith("Documento:"):
                current_documento = sv.replace("Documento:", "").strip()
                break

        # Detectar "Grupo:"
        for v in vals.values():
            sv = str(v).strip()
            if sv.startswith("Grupo:"):
                # strip() extra para eliminar espacios, tabs y saltos de línea ocultos
                current_grupo = " ".join(sv.replace("Grupo:", "").split())
                break

        # Detectar fila de encabezados de columnas (contiene 'Fecha' y 'TOTAL')
        row_str_vals = [str(v).strip() for v in vals.values()]
        if "Fecha" in row_str_vals and "TOTAL" in row_str_vals:
            col_map = {}
            for idx, v in vals.items():
                if v is not None:
                    name = str(v).strip().replace("\n", " ")
                    col_map[idx] = name
            continue

        # Detectar fila de datos (primera columna no-nula es un día de la semana)
        if col_map and current_nombre:
            first_val = None
            first_idx = None
            for i in sorted(vals.keys()):
                if vals[i] is not None:
                    first_val = str(vals[i]).strip().lower()
                    first_idx = i
                    break

            if first_val and first_val in DIAS:
                # Es una fila de datos
                rec = {
                    "Nombre": current_nombre,
                    "Documento": current_documento,
                    "Grupo": current_grupo,
                }
                # Mapear columnas según los encabezados dinámicos
                for col_idx, col_name in col_map.items():
                    raw = vals.get(col_idx)
                    if raw is not None and str(raw).strip() not in ("", " "):
                        rec[col_name] = str(raw).strip()
                    else:
                        rec[col_name] = None

                # Normalizar campo Día
                rec["Día"] = str(vals.get(first_idx, "")).strip().capitalize()

                records.append(rec)

    if not records:
        return pd.DataFrame()

    df = pd.DataFrame(records)

    # ── Normalizar fecha ──
    if "Fecha" in df.columns:
        df["Fecha"] = pd.to_datetime(
            df["Fecha"].str.strip(), errors="coerce",
            format="%m/%d/%Y",
        )
        # Intentar formato alternativo si falló
        mask = df["Fecha"].isna()
        if mask.any():
            df.loc[mask, "Fecha"] = pd.to_datetime(
                df.loc[mask, "Fecha_raw"] if "Fecha_raw" in df.columns else df.loc[mask, "Fecha"],
                errors="coerce",
            )

    # ── Semana y mes ──
    if "Fecha" in df.columns:
        df["Semana"] = df["Fecha"].dt.isocalendar().week.astype("Int64")
        df["Semana_etiqueta"] = df["Fecha"].dt.to_period("W").astype(str)
        df["Mes"] = df["Fecha"].dt.strftime("%B %Y")
        df["Mes_num"] = df["Fecha"].dt.month

    # ── Convertir columnas de horas a float ──
    hour_candidates = list(HORA_COLS.keys()) + ["TOTAL"]
    for col in hour_candidates:
        if col in df.columns:
            df[col] = pd.to_numeric(df[col], errors="coerce").fillna(0.0)

    # Asegurar TOTAL: si la columna se llama 'TOTAL' o coincide parcialmente
    total_col = next((c for c in df.columns if c.upper() == "TOTAL"), None)
    if total_col and total_col != "TOTAL":
        df["TOTAL"] = df[total_col]

    # Filtrar filas sin fecha válida
    if "Fecha" in df.columns:
        df = df[df["Fecha"].notna()].copy()

    return df


# ─────────────────────────────────────────────
# SQL HISTORICO (ADITIVO, NO REEMPLAZA EXCEL)
# ─────────────────────────────────────────────


def _get_secret_or_env(key: str, default=None):
    try:
        if key in st.secrets:
            return st.secrets[key]
    except Exception:
        pass
    return os.getenv(key, default)


def _db_cfg() -> dict:
    return {
        "server": _get_secret_or_env("DB_SERVER"),
        "port": int(_get_secret_or_env("DB_PORT", "1433")),
        "name": _get_secret_or_env("DB_NAME"),
        "user": _get_secret_or_env("DB_USER"),
        "password": _get_secret_or_env("DB_PASSWORD"),
    }


def _db_missing(cfg: dict) -> list:
    missing = []
    if not cfg.get("server"):
        missing.append("DB_SERVER")
    if not cfg.get("name"):
        missing.append("DB_NAME")
    if not cfg.get("user"):
        missing.append("DB_USER")
    if not cfg.get("password"):
        missing.append("DB_PASSWORD")
    return missing


def _normalize_doc_series(s: pd.Series) -> pd.Series:
    return (
        s.astype(str)
        .str.strip()
        .str.upper()
        .str.replace(r"\.0$", "", regex=True)
        .str.replace(r"[^0-9A-Z]", "", regex=True)
    )


def _to_bool(value, default: bool = True) -> bool:
    if value is None:
        return default
    if isinstance(value, bool):
        return value
    return str(value).strip().lower() in {"1", "true", "yes", "y", "si", "on"}


def _sql_connect(cfg: dict):
    validate_host = _to_bool(_get_secret_or_env("DB_VALIDATE_HOST", True), True)
    return pytds.connect(
        server=cfg["server"],
        port=cfg["port"],
        database=cfg["name"],
        user=cfg["user"],
        password=cfg["password"],
        validate_host=validate_host,
    )


@st.cache_data(show_spinner=False)
def _sql_test_connection(cfg: dict) -> bool:
    conn = _sql_connect(cfg)
    cur = conn.cursor()
    cur.execute("SELECT 1")
    ok = cur.fetchone()[0] == 1
    conn.close()
    return ok


def _classify_sql_concept(concept_id: str, concept_desc: str, concept_abbr: str) -> str:
    abbr = str(concept_abbr or "").upper().strip()
    txt = " ".join(str(x or "") for x in [concept_id, concept_desc, concept_abbr]).upper()
    txt = " ".join(txt.split())

    # Mapeo estricto por abreviatura para acercar SQL al reporte operativo.
    if abbr in {"SALBAS"}:
        return "DO"
    if abbr in {"REC NOCT", "RNO"}:
        return "RNO"
    if abbr in {"HED", "HEDO"}:
        return "HEDO"
    if abbr in {"HEN", "HENO"}:
        return "HENO"
    if abbr in {"RNF", "RNFS"}:
        return "RNF"
    if abbr in {"HEDF"}:
        return "HEDF"
    if abbr in {"HENF"}:
        return "HENF"
    if abbr in {"HEDN", "RNDOM"}:
        return "RNDOM"
    if abbr in {"HEDD", "RDOM"}:
        return "RDOM"
    if abbr in {"ODOM"}:
        return "ODOM"
    if abbr in {"FEST"}:
        return "FEST"
    if abbr in {"OFEST"}:
        return "OFEST"
    if abbr in {"ORNF"}:
        return "ORNF"
    # Conceptos compensados/licencias/vacaciones no son comparables con el archivo.
    if abbr in {"DOMCOMPE", "VAC", "VAC ", "CUOTA S", "LIC MATER", "LIC DFLIA", "LIC CAL DOM", "LIC N REM", "INCAP G 66%", "INCAP ACCI T", "HECAPAC"}:
        return "OTRAS"

    if "RNDOM" in txt or ("RECARGO" in txt and "NOCT" in txt and ("DOM" in txt or "DOMINICAL" in txt)):
        return "RNDOM"
    if "RDOM" in txt or ("RECARGO" in txt and ("DOM" in txt or "DOMINICAL" in txt)):
        return "RDOM"
    if "RNF" in txt or ("RECARGO" in txt and "NOCT" in txt and "FEST" in txt):
        return "RNF"
    if "HEDF" in txt or ("HORA EXTRA DIURNA" in txt and "FEST" in txt):
        return "HEDF"
    if "HENF" in txt or ("HORA EXTRA NOCTURNA" in txt and "FEST" in txt):
        return "HENF"
    if "OFEST" in txt:
        return "OFEST"
    if "RECARGO NOCTURNO" in txt and "DOM" not in txt and "FEST" not in txt:
        return "RNO"
    if "HORA EXTRA DIURNA" in txt and "DOM" not in txt and "FEST" not in txt:
        return "HEDO"
    if "HORA EXTRA NOCTURNA" in txt and "DOM" not in txt and "FEST" not in txt:
        return "HENO"
    if "DOMING" in txt or "DOMINICAL" in txt:
        return "DOM"
    if "FEST" in txt:
        return "FEST"
    if "SALARIO BASICO" in txt or "JORNADA" in txt:
        return "DO"
    return "OTRAS"


@st.cache_data(show_spinner="Consultando horas historicas en SQL...")
def _sql_fetch_hours(cfg: dict, fecha_ini, fecha_fin) -> pd.DataFrame:
    conn = _sql_connect(cfg)
    cur = conn.cursor()
    query = """
    SELECT
        CAST(m.c0602_fecha AS date) AS Fecha,
        t.f200_nit AS Documento,
        LTRIM(RTRIM(CONCAT(ISNULL(t.f200_nombres, ''), ' ', ISNULL(t.f200_apellido1, ''), ' ', ISNULL(t.f200_apellido2, '')))) AS Nombre,
        c.c0501_id AS ConceptoID,
        c.c0501_descripcion AS Concepto,
        c.c0501_abreviatura AS Abreviatura,
        SUM(ISNULL(m.c0602_horas, 0)) AS Horas
    FROM dbo.w0602_movto_nomina m
    INNER JOIN dbo.w0501_conceptos c ON c.c0501_rowid = m.c0602_rowid_concepto
    LEFT JOIN dbo.t200_mm_terceros t ON t.f200_rowid = m.c0602_rowid_tercero AND t.f200_id_cia = m.c0602_id_cia
    WHERE m.c0602_horas IS NOT NULL
      AND m.c0602_horas > 0
      AND CAST(m.c0602_fecha AS date) BETWEEN %s AND %s
    GROUP BY
        CAST(m.c0602_fecha AS date),
        t.f200_nit,
        t.f200_nombres,
        t.f200_apellido1,
        t.f200_apellido2,
        c.c0501_id,
        c.c0501_descripcion,
        c.c0501_abreviatura
    """
    cur.execute(query, (fecha_ini, fecha_fin))
    rows = cur.fetchall()
    conn.close()

    raw = pd.DataFrame(rows, columns=["Fecha", "Documento", "Nombre", "ConceptoID", "Concepto", "Abreviatura", "Horas"])
    if raw.empty:
        return raw

    raw["Fecha"] = pd.to_datetime(raw["Fecha"])
    raw["Categoria"] = raw.apply(
        lambda r: _classify_sql_concept(r["ConceptoID"], r["Concepto"], r["Abreviatura"]),
        axis=1,
    )
    # NOTA: SALBAS en SQL no almacena horas por turno sino el salario base acumulado.
    # Reclasificarlo por día de semana inflaría DOM con valores monetarios, NO con horas reales.
    # DOM en SQL solo se alimenta de conceptos explícitamente dominicales (ej. RDOM, RNDOM).
    # La diferencia residual entre DOM SQL y DOM Excel es una limitación estructural del ERP.

    out = (
        raw.groupby(["Fecha", "Documento", "Nombre", "Categoria"], as_index=False)["Horas"]
        .sum()
        .pivot_table(index=["Fecha", "Documento", "Nombre"], columns="Categoria", values="Horas", fill_value=0)
        .reset_index()
    )
    out.columns = [str(c) for c in out.columns]
    for col in ["DO", "RNO", "HEDO", "HENO", "DOM", "RNF", "HEDF", "HENF",
                "FEST", "RNDOM", "RDOM", "ODOM", "ORNF", "OFEST", "OTRAS"]:
        if col not in out.columns:
            out[col] = 0.0
    out["TOTAL"] = out[["DO", "RNO", "HEDO", "HENO", "DOM", "RNF", "HEDF", "HENF",
                         "FEST", "RNDOM", "RDOM", "ODOM", "ORNF", "OFEST"]].sum(axis=1)
    out["Mes"] = out["Fecha"].dt.to_period("M").astype(str)
    return out


# ─────────────────────────────────────────────
# CARGA DE DATOS
# ─────────────────────────────────────────────

if uploaded_file is None:
    st.info("📂 Sube un archivo Excel de Siesa Access desde el panel izquierdo para comenzar.")
    st.stop()

try:
    df = parse_excel(excel_path)
except Exception as e:
    st.error(f"No se pudieron extraer datos del archivo. Verifica el formato.\n\nError: {e}")
    st.stop()

if df.empty:
    st.error("No se pudieron extraer datos del archivo. Verifica el formato.")
    st.stop()

# Normalizar columna Grupo: eliminar espacios redundantes
if "Grupo" in df.columns:
    df["Grupo"] = df["Grupo"].astype(str).str.strip().str.replace(r"\s+", " ", regex=True)
    df["Grupo"] = df["Grupo"].replace("None", "(Sin grupo)").fillna("(Sin grupo)")

# Columnas de horas disponibles en el DataFrame
HORAS_DISPONIBLES = [c for c in HORA_COLS if c in df.columns]
if "TOTAL" in df.columns and "TOTAL" not in HORAS_DISPONIBLES:
    HORAS_DISPONIBLES.append("TOTAL")

# ── Detectar cambio de archivo y resetear filtros de tiendas/empleados ──
_current_file_id = uploaded_file.name if uploaded_file else ""
if st.session_state.get("_last_file_id") != _current_file_id:
    st.session_state["_last_file_id"] = _current_file_id
    for _k in ["sel_grupos_ms", "sel_personas_ms", "_prev_grupos"]:
        st.session_state.pop(_k, None)

# ─────────────────────────────────────────────
# SIDEBAR – FILTROS (continuación)
# ─────────────────────────────────────────────
    # CSS para limitar altura de los multiselect y hacerlos scrolleables
    st.markdown(
        """
        <style>
        /* Área de tags seleccionados – max-height + scroll */
        section[data-testid="stSidebar"] div[data-baseweb="select"] > div:first-child {
            max-height: 90px;
            overflow-y: auto;
        }
        </style>
        """,
        unsafe_allow_html=True,
    )

with st.sidebar:
    # SECCIÓN 2: FILTRO POR TIENDA / GRUPO  (primero tienda, luego empleados)
    st.markdown("### 🏪 Filtros - Tiendas")
    _grupos_raw = sorted([g for g in df["Grupo"].fillna("(Sin grupo)").unique()])

    # Inicializar/validar session_state para tiendas
    if "sel_grupos_ms" not in st.session_state:
        # Primera carga o cambio de archivo: seleccionar todas
        st.session_state["sel_grupos_ms"] = _grupos_raw
    else:
        # Filtrar solo grupos que existen en el archivo actual; si el usuario
        # limpió la selección (lista vacía), se respeta sin forzar a todas.
        st.session_state["sel_grupos_ms"] = [
            g for g in st.session_state["sel_grupos_ms"] if g in _grupos_raw
        ]

    # Botones Seleccionar todo / Limpiar
    _bt1, _bt2 = st.columns(2)
    with _bt1:
        if st.button("✅ Todas", key="btn_all_grupos", use_container_width=True):
            st.session_state["sel_grupos_ms"] = _grupos_raw
            st.rerun()
    with _bt2:
        if st.button("🗑 Ninguna", key="btn_none_grupos", use_container_width=True):
            st.session_state["sel_grupos_ms"] = []
            st.rerun()

    sel_grupos = st.multiselect(
        "Tienda(s) / Grupo(s)",
        _grupos_raw,
        key="sel_grupos_ms",
        help="Selecciona una o varias tiendas · usa los botones para seleccionar/limpiar todo",
    )

    st.divider()

    # SECCIÓN 2b: FILTROS PRINCIPALES – empleados de las tiendas seleccionadas
    st.markdown("### 🔎 Filtros - Empleados")
    _df_por_tienda = df[df["Grupo"].fillna("(Sin grupo)").isin(sel_grupos)]
    personas = sorted(_df_por_tienda["Nombre"].unique())

    # Resetear empleados si cambió la selección de tiendas
    _grupos_cambiaron = set(st.session_state.get("_prev_grupos", [])) != set(sel_grupos)
    if "sel_personas_ms" not in st.session_state or _grupos_cambiaron:
        st.session_state["sel_personas_ms"] = personas
    else:
        # Mantener solo empleados válidos; respetar lista vacía si el usuario la limpió
        st.session_state["sel_personas_ms"] = [
            p for p in st.session_state["sel_personas_ms"] if p in personas
        ]
    st.session_state["_prev_grupos"] = sel_grupos

    _be1, _be2 = st.columns(2)
    with _be1:
        if st.button("✅ Todos", key="btn_all_emp", use_container_width=True):
            st.session_state["sel_personas_ms"] = personas
            st.rerun()
    with _be2:
        if st.button("🗑 Ninguno", key="btn_none_emp", use_container_width=True):
            st.session_state["sel_personas_ms"] = []
            st.rerun()

    sel_personas = st.multiselect(
        "Empleado(s)",
        personas,
        key="sel_personas_ms",
        help="Lista de empleados de las tiendas seleccionadas arriba",
    )

    st.divider()

    # SECCIÓN 3: PERÍODO DE ANÁLISIS
    st.markdown("### <i class='lucide lucide-calendar'></i> Período y vista", unsafe_allow_html=True)
    
    # Período de análisis
    vista = st.radio(
        "Agrupar por",
        ["Día", "Semana", "Mes"],
        index=0,
    )

    st.divider()

    # SECCIÓN 4: RANGO DE FECHAS
    st.markdown("### <i class='lucide lucide-calendar-range'></i> Rango de fechas", unsafe_allow_html=True)

    st.divider()

    # Rango de fechas
    min_fecha = df["Fecha"].min().date()
    max_fecha = df["Fecha"].max().date()
    rango = st.date_input(
        "Desde — Hasta",
        value=(min_fecha, max_fecha),
        min_value=min_fecha,
        max_value=max_fecha,
    )
    if isinstance(rango, (list, tuple)) and len(rango) == 2:
        fecha_ini, fecha_fin = rango
    else:
        fecha_ini, fecha_fin = min_fecha, max_fecha

    st.divider()

    # SECCIÓN 5: MÉTRICA DE HORAS
    st.markdown("### <i class='lucide lucide-activity'></i> Métrica de horas", unsafe_allow_html=True)
    
    # Tipo de hora
    tipo_hora = st.selectbox(
        "Selecciona métrica",
        HORAS_DISPONIBLES,
        index=HORAS_DISPONIBLES.index("TOTAL") if "TOTAL" in HORAS_DISPONIBLES else 0,
        format_func=lambda x: f"{x} – {HORA_COLS.get(x, '')}",
    )



# ─────────────────────────────────────────────
# APLICAR FILTROS
# ─────────────────────────────────────────────
dff = df[
    df["Nombre"].isin(sel_personas)
    & df["Grupo"].fillna("(Sin grupo)").isin(sel_grupos)
    & (df["Fecha"].dt.date >= fecha_ini)
    & (df["Fecha"].dt.date <= fecha_fin)
].copy()
dff["Grupo"] = dff["Grupo"].fillna("(Sin grupo)")

# ─────────────────────────────────────────────
# ENCABEZADO
# ─────────────────────────────────────────────
st.title("FYC Calzado – Reporte de Labor")
st.caption("Fuente: Siesa Access · Octubre 2025")

if dff.empty:
    st.warning("No hay datos para los filtros seleccionados.")
    st.stop()

st.markdown(
    """
    <style>
    /* Navbar de tabs fija en la parte superior del contenido */
    div[data-testid="stTabs"] > div[role="tablist"] {
        position: sticky;
        top: 0;
        z-index: 999;
        background: #F7F5F3;
        border-bottom: 1px solid #E4DDD7;
        padding-top: 0.35rem;
        padding-bottom: 0.25rem;
    }
    </style>
    """,
    unsafe_allow_html=True,
)

tabs = st.tabs([
    "Por día",
    "Por semana",
    "Por mes",
    "Detalle completo",
    "⚖️ Cumplimiento Ley 2466",
    "Total laborado",
    "🏪 Por Tienda",
    "🔌 SQL Histórico",
])

# Redondear todas las columnas numéricas de horas a máximo 2 decimales
_hora_num_cols = [c for c in list(HORA_COLS.keys()) + ["TOTAL"] if c in dff.columns]
for _c in _hora_num_cols:
    dff[_c] = dff[_c].round(2)

# ─────────────────────────────────────────────
# KPIs
# ─────────────────────────────────────────────
def _col_sum(col): return dff[col].sum() if col in dff.columns else 0

total_horas       = dff["TOTAL"].sum() if "TOTAL" in dff.columns else 0
total_do          = _col_sum("DO")
total_rno         = _col_sum("RNO")
total_hedo        = _col_sum("HEDO")
total_heno        = _col_sum("HENO")
total_dom         = _col_sum("DOM")
total_rnf         = _col_sum("RNF")
total_hedf        = _col_sum("HEDF")
total_henf        = _col_sum("HENF")
total_fest        = _col_sum("FEST")
total_rndom       = _col_sum("RNDOM")
total_rdom        = _col_sum("RDOM")
total_extras_ord  = total_hedo + total_heno
total_extras_fest = total_hedf + total_henf
dia_persona_laborados = int((dff["DO"] > 0).sum()) if "DO" in dff.columns else len(dff)
dias_calendario_periodo = (pd.to_datetime(fecha_fin) - pd.to_datetime(fecha_ini)).days + 1
# ── Fila 1: resumen principal ──
st.markdown("##### Resumen general")
k1, k2, k3, k4 = st.columns(4)
k1.metric("✅ Total horas", f"{total_horas:,.2f}")
k2.metric("DO – Jornada ordinaria", f"{total_do:,.2f}")
k3.metric("Jornadas (persona-día)", dia_persona_laborados, help="Conteo de registros con DO > 0. No son días calendario del mes.")
k4.metric("Días calendario período", dias_calendario_periodo)

# ── Fila 2: horas extras separadas (cada tipo con valor de nómina distinto) ──
st.markdown("##### Horas extras – desglose por tipo de pago")
e1, e2, e3, e4 = st.columns(4)
e1.metric("HEDO – H. Extra diurna ord.",   f"{total_hedo:,.2f}", help="Hora extra diurna en día ordinario")
e2.metric("HENO – H. Extra nocturna ord.", f"{total_heno:,.2f}", help="Hora extra nocturna en día ordinario")
e3.metric("HEDF – H. Extra diurna fest.",  f"{total_hedf:,.2f}", help="Hora extra diurna en día festivo")
e4.metric("HENF – H. Extra noct. fest.",   f"{total_henf:,.2f}", help="Hora extra nocturna en día festivo")

# ── Fila 3: otros tipos de hora ──
st.markdown("##### Otros tipos de hora")
d1, d2, d3, d4, d5 = st.columns(5)
d1.metric("RNO",       f"{total_rno:,.2f}",            help="Recargo nocturno ord.")
d2.metric("DOM",       f"{total_dom:,.2f}",            help="Dominical")
d3.metric("RNF",       f"{total_rnf:,.2f}",            help="Rec. noct. festivo")
d4.metric("FEST",      f"{total_fest:,.2f}",           help="Festivo")
d5.metric("RNDOM/RDOM",f"{(total_rndom + total_rdom):,.2f}", help="Rec. noct. dom. + Rec. dom.")

# ── Verificación de sumatoria ──
# JORNADA es columna de referencia del turno, NO es componente del TOTAL en Siesa.
# Incluirla duplicaría las horas ordinarias (DO ya las cubre).
_cols_verificar = ["DO", "RNO", "HEDO", "HENO", "DOM", "RNF", "HEDF", "HENF",
                   "FEST", "RNDOM", "RDOM", "ODOM", "ORNF", "OEDF", "OFEST"]
_suma_parciales = sum(_col_sum(c) for c in _cols_verificar)
_diff = abs(total_horas - _suma_parciales)
if total_horas > 0 and _diff > 0.5:
    st.warning(
        f"⚠️ Diferencia de {_diff:,.2f} h entre la suma de columnas parciales "
        f"({_suma_parciales:,.2f} h) y el TOTAL del archivo ({total_horas:,.2f} h). "
        f"Verifique si el archivo tiene columnas adicionales no reconocidas."
    )
else:
    st.success(f"✔ Verificación OK: suma parciales = {_suma_parciales:,.2f} h  ═  TOTAL archivo = {total_horas:,.2f} h")

st.divider()

# ── Tarjeta de detalle por empleado (si se filtra uno solo) ──
if len(sel_personas) == 1:
    emp_name = sel_personas[0]
    st.markdown(f"### 👤 Detalle – {emp_name}")

    _detail_cols = [c for c in ["DO", "RNO", "HEDO", "HENO", "DOM", "RNF",
                                 "HEDF", "HENF", "FEST", "RNDOM", "RDOM",
                                 "ODOM", "ORNF", "OEDF", "OFEST", "TOTAL"]
                    if c in dff.columns]

    _emp_totals = {c: dff[c].sum() for c in _detail_cols}

    # Construir tabla de desglose
    _filas = []
    _grupo_map = {
        "DO  – Jornada ordinaria":          ["DO"],
        "RNO – Recargo nocturno ord.":       ["RNO"],
        "HEDO – H. Extra diurna ord.":       ["HEDO"],
        "HENO – H. Extra nocturna ord.":     ["HENO"],
        "DOM  – Dominical":                  ["DOM"],
        "FEST – Festivo":                    ["FEST"],
        "RNF  – Rec. noct. festivo":         ["RNF"],
        "HEDF – H. Extra diurna fest.":      ["HEDF"],
        "HENF – H. Extra noct. fest.":       ["HENF"],
        "RNDOM – Rec. noct. dom.":           ["RNDOM"],
        "RDOM  – Rec. dom.":                 ["RDOM"],
        "Otros recargos":                    ["ODOM", "ORNF", "OEDF", "OFEST"],
    }
    for grupo, cols_g in _grupo_map.items():
        subtotal = sum(_emp_totals.get(c, 0) for c in cols_g)
        if subtotal > 0:
            _filas.append({
                "Grupo": grupo,
                "Columnas": " + ".join(c for c in cols_g if c in dff.columns),
                "Total horas": round(subtotal, 2),
            })
    _total_row = {"Grupo": "🔢 TOTAL GENERAL", "Columnas": "TOTAL",
                  "Total horas": round(_emp_totals.get("TOTAL", 0), 2)}
    _suma_row  = {"Grupo": "∑ Suma parciales", "Columnas": " + ".join(
        c for c in _detail_cols if c != "TOTAL"),
        "Total horas": round(sum(v for k, v in _emp_totals.items() if k != "TOTAL"), 2)}

    _df_detalle = pd.DataFrame(_filas + [_suma_row, _total_row])

    def _color_detalle(row):
        if row["Grupo"] == "🔢 TOTAL GENERAL":
            return ["background-color:#8A2F1F; color:white; font-weight:bold"] * len(row)
        if row["Grupo"] == "∑ Suma parciales":
            return ["background-color:#FFF3F0; color:#3D1A0E; font-weight:bold"] * len(row)
        return [""] * len(row)

    st.dataframe(
        _df_detalle.style
            .apply(_color_detalle, axis=1)
            .format({"Total horas": "{:.2f}"}),
        use_container_width=True,
        hide_index=True,
        height=min(80 + len(_df_detalle) * 40, 500),
    )
    _xlsx_detalle_emp = df_to_excel_grouped(_df_detalle, f"Detalle {emp_name}", group_col="_none_")
    st.download_button(
        "⬇ Descargar Excel (detalle empleado)",
        data=_xlsx_detalle_emp,
        file_name=f"detalle_{emp_name.replace(' ','_')}.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        key="dl_detalle_emp",
    )
    st.divider()

# ─────────────────────────────────────────────
# GRÁFICA PRINCIPAL – según vista seleccionada
# ─────────────────────────────────────────────
st.subheader(f"Horas por empleado · agrupado por {vista.lower()}")

if vista == "Día":
    grp_col = "Fecha"
    grp_label = "Fecha"
    grp_df = dff.groupby(["Nombre", "Fecha"])[tipo_hora].sum().reset_index()
    grp_df["Fecha_str"] = grp_df["Fecha"].dt.strftime("%d/%m")
    fig = px.bar(
        grp_df,
        x="Fecha_str",
        y=tipo_hora,
        color="Nombre",
        barmode="group",
        labels={"Fecha_str": "Día", tipo_hora: "Horas"},
        height=420,
    )
    fig.update_layout(xaxis_title="Día", legend_title="Empleado")

elif vista == "Semana":
    grp_df = dff.groupby(["Nombre", "Semana_etiqueta"])[tipo_hora].sum().reset_index()
    fig = px.bar(
        grp_df,
        x="Semana_etiqueta",
        y=tipo_hora,
        color="Nombre",
        barmode="group",
        labels={"Semana_etiqueta": "Semana", tipo_hora: "Horas"},
        height=420,
    )
    fig.update_layout(xaxis_title="Semana (ISO)", legend_title="Empleado")

else:  # Mes
    grp_df = dff.groupby(["Nombre", "Mes"])[tipo_hora].sum().reset_index()
    fig = px.bar(
        grp_df,
        x="Mes",
        y=tipo_hora,
        color="Nombre",
        barmode="group",
        labels={"Mes": "Mes", tipo_hora: "Horas"},
        height=420,
    )
    fig.update_layout(xaxis_title="Mes", legend_title="Empleado")

fig.update_layout(
    plot_bgcolor="rgba(0,0,0,0)",
    paper_bgcolor="rgba(0,0,0,0)",
    font_size=12,
)
import streamlit.components.v1 as components
from plotly.io import to_html
from streamlit_js_eval import streamlit_js_eval

# --- Detección automática de móvil/tablet usando ancho real del navegador ---
_screen_width = streamlit_js_eval(js_expressions="window.innerWidth", key="_screen_w")
is_mobile = (_screen_width is not None and int(_screen_width) < 700)

# TOP_N empleados en móvil ---
TOP_N = 5

# Definir fig_height ANTES de cualquier uso
fig_height = 420
if is_mobile:
    fig_height = 260

# --- Filtrar top N empleados en móvil ANTES de crear la figura ---
TOP_N = 5
if is_mobile and 'Nombre' in grp_df.columns:
    if 'Semana_etiqueta' in grp_df.columns:
        top_empleados = grp_df.groupby('Nombre')[tipo_hora].sum().nlargest(TOP_N).index
        grp_df = grp_df[grp_df['Nombre'].isin(top_empleados)]
    elif 'Mes' in grp_df.columns:
        top_empleados = grp_df.groupby('Nombre')[tipo_hora].sum().nlargest(TOP_N).index
        grp_df = grp_df[grp_df['Nombre'].isin(top_empleados)]



import numpy as np
# Regenerar la figura con el df filtrado (o el original si no es móvil)
if is_mobile:
    # Gráfica de dona en móvil
    if 'Nombre' in grp_df.columns and tipo_hora in grp_df.columns:
        pie_df = grp_df.groupby('Nombre')[tipo_hora].sum().reset_index()
        pie_df = pie_df[pie_df[tipo_hora] > 0]
        fig = px.pie(
            pie_df,
            values=tipo_hora,
            names="Nombre",
            hole=0.45,
            height=fig_height,
            title=f"Distribución de {tipo_hora} por empleado",
        )
        fig.update_traces(
            textinfo='percent',
            textposition='inside',
            insidetextorientation='radial',
        )
        fig.update_layout(
            legend_title="Empleado",
            font_size=12,
            showlegend=True,
            legend=dict(orientation='v', x=1.02, y=0.5),
        )
    else:
        # Fallback: gráfico vacío
        fig = go.Figure()
        fig.add_annotation(text="Sin datos para mostrar", showarrow=False)
else:
    # Barras verticales en desktop
    if 'Semana_etiqueta' in grp_df.columns:
        fig = px.bar(
            grp_df,
            x="Semana_etiqueta",
            y=tipo_hora,
            color="Nombre",
            barmode="group",
            height=fig_height,
            labels={"Semana_etiqueta": "Semana", tipo_hora: "Horas"},
        )
        fig.update_layout(xaxis_title="Semana (ISO)", legend_title="Empleado")
    elif 'Mes' in grp_df.columns:
        fig = px.bar(
            grp_df,
            x="Mes",
            y=tipo_hora,
            color="Nombre",
            barmode="group",
            height=fig_height,
            labels={"Mes": "Mes", tipo_hora: "Horas"},
        )
        fig.update_layout(xaxis_title="Mes", legend_title="Empleado")
    elif 'Fecha_str' in grp_df.columns:
        fig = px.bar(
            grp_df,
            x="Fecha_str",
            y=tipo_hora,
            color="Nombre",
            barmode="group",
            height=fig_height,
            labels={"Fecha_str": "Día", tipo_hora: "Horas"},
        )
        fig.update_layout(xaxis_title="Día", legend_title="Empleado")

# Ajustes visuales extra para móvil
if is_mobile:
    fig.update_layout(
        font_size=11,
        legend_font_size=10,
        legend_itemwidth=40,
        margin=dict(l=10, r=10, t=40, b=40),
    )
    st.info(f"Mostrando solo los {TOP_N} empleados con más horas para mejor visualización en móvil.")
else:
    fig.update_layout(font_size=12)
fig.update_layout(height=fig_height)
st.plotly_chart(fig, use_container_width=True)

# ─────────────────────────────────────────────
# FUNCIONES DE EXPORTACIÓN
# ─────────────────────────────────────────────
import io
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from fpdf import FPDF

# Columnas de horas extras reconocidas
COLS_HORAS_EXTRA = ["HEDO", "HENO", "HEDF", "HENF"]

# Límites legales (Ley 2466 de 2025 / Código Sustantivo del Trabajo)
LIMITE_DIARIO_HE = 2.0    # máximo 2 horas extras por día
LIMITE_SEMANAL_HE = 12.0  # máximo 12 horas extras por semana


def calcular_cumplimiento(df: pd.DataFrame) -> tuple[pd.DataFrame, pd.DataFrame]:
    """
    Verifica el cumplimiento de los límites legales de horas extras.

    Norma aplicada: Ley 2466 de 2025 / Código Sustantivo del Trabajo.
    - Máximo 2 horas extras diarias por empleado.
    - Máximo 12 horas extras semanales por empleado.

    Parámetros
    ----------
    df : pd.DataFrame
        DataFrame con columnas de horas extras, Nombre, Fecha, Semana_etiqueta.

    Retorna
    -------
    (df_diario, df_semanal) : tuple de DataFrames
        df_diario  → una fila por empleado/día con total extras y estado.
        df_semanal → una fila por empleado/semana con total extras y estado.
        En ambos, la columna "Estado" indica "✅ OK" o "🚨 EXCEDE LÍMITE".
    """
    # Columnas de horas extras que existen en el DataFrame
    extras_cols = [c for c in COLS_HORAS_EXTRA if c in df.columns]

    work = df.copy()
    work["_HE_Total"] = work[extras_cols].sum(axis=1) if extras_cols else 0.0

    # ── Resumen diario ──
    grp_dia = (
        work.groupby(["Nombre", "Fecha"])["_HE_Total"]
        .sum()
        .reset_index()
        .rename(columns={"_HE_Total": "H. Extra del día"})
    )
    grp_dia["Fecha"] = grp_dia["Fecha"].dt.strftime("%d/%m/%Y")
    grp_dia["Límite diario"] = LIMITE_DIARIO_HE
    grp_dia["Exceso diario"] = (grp_dia["H. Extra del día"] - LIMITE_DIARIO_HE).clip(lower=0).round(2)
    grp_dia["Estado"] = grp_dia["H. Extra del día"].apply(
        lambda x: "🚨 EXCEDE LÍMITE" if x > LIMITE_DIARIO_HE else "✅ OK"
    )
    # Excluir días sin horas extras (no aplica verificación)
    grp_dia = grp_dia[grp_dia["H. Extra del día"] > 0].reset_index(drop=True)

    # ── Resumen semanal ──
    if "Semana_etiqueta" in work.columns:
        grp_sem = (
            work.groupby(["Nombre", "Semana_etiqueta"])["_HE_Total"]
            .sum()
            .reset_index()
            .rename(columns={"Semana_etiqueta": "Semana", "_HE_Total": "H. Extra semana"})
        )
    else:
        grp_sem = pd.DataFrame(columns=["Nombre", "Semana", "H. Extra semana"])

    grp_sem["Límite semanal"] = LIMITE_SEMANAL_HE
    grp_sem["Exceso semanal"] = (grp_sem["H. Extra semana"] - LIMITE_SEMANAL_HE).clip(lower=0).round(2)
    grp_sem["Estado"] = grp_sem["H. Extra semana"].apply(
        lambda x: "🚨 EXCEDE LÍMITE" if x > LIMITE_SEMANAL_HE else "✅ OK"
    )
    # Excluir semanas sin horas extras (no aplica verificación)
    grp_sem = grp_sem[grp_sem["H. Extra semana"] > 0].reset_index(drop=True)

    return grp_dia, grp_sem


def excel_cumplimiento(df_diario: pd.DataFrame, df_semanal: pd.DataFrame) -> bytes:
    """
    Genera un Excel con dos hojas: incumplimientos diarios y semanales.
    Las filas con exceso se resaltan en rojo para fácil identificación.

    Retorna
    -------
    bytes
        Contenido binario del .xlsx.
    """
    wb = Workbook()

    COLOR_OK       = "E8F5E9"  # verde claro
    COLOR_EXCEDE   = "FFEBEE"  # rojo claro
    COLOR_HEADER   = "8A2F1F"  # rojo corporativo
    thin = Side(style="thin", color="CCCCCC")
    brd  = Border(left=thin, right=thin, top=thin, bottom=thin)

    def _write_sheet(ws, df, titulo):
        # Fila de título
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(df.columns))
        title_cell = ws.cell(row=1, column=1, value=titulo)
        title_cell.font = Font(bold=True, size=12, color="FFFFFF")
        title_cell.fill = PatternFill("solid", fgColor=COLOR_HEADER)
        title_cell.alignment = Alignment(horizontal="center")

        # Encabezados
        hdr_font = Font(bold=True, color="FFFFFF", size=10)
        hdr_fill = PatternFill("solid", fgColor="5D2010")
        for ci, col in enumerate(df.columns, 1):
            c = ws.cell(row=2, column=ci, value=col)
            c.font = hdr_font
            c.fill = hdr_fill
            c.alignment = Alignment(horizontal="center")
            c.border = brd
        ws.freeze_panes = "A3"

        # Filas de datos
        for ri, (_, row) in enumerate(df.iterrows(), 3):
            es_exceso = "EXCEDE" in str(row.get("Estado", ""))
            fill = PatternFill("solid", fgColor=COLOR_EXCEDE if es_exceso else COLOR_OK)
            for ci, val in enumerate(row, 1):
                c = ws.cell(row=ri, column=ci, value=val)
                c.fill = fill
                c.border = brd
                c.font = Font(size=9, bold=es_exceso, color="B71C1C" if es_exceso else "1B5E20")
                if isinstance(val, float):
                    c.number_format = "0.00"

        # Auto-ancho
        for ci, col in enumerate(df.columns, 1):
            max_len = max(len(str(col)), *(len(str(v)) for v in df.iloc[:, ci - 1]))
            ws.column_dimensions[get_column_letter(ci)].width = min(max_len + 2, 35)

    # Hoja 1 – diario
    ws1 = wb.active
    ws1.title = "Incumplimiento Diario"
    _write_sheet(ws1, df_diario, "Reporte de cumplimiento – Horas extras diarias (máx. 2 h/día)")

    # Hoja 2 – semanal
    ws2 = wb.create_sheet("Incumplimiento Semanal")
    _write_sheet(ws2, df_semanal, "Reporte de cumplimiento – Horas extras semanales (máx. 12 h/semana)")

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()


def df_to_excel_grouped(df, title, group_col="Nombre"):
    """
    Genera un archivo Excel (.xlsx) estructurado y visualmente organizado.

    El archivo tiene:
    - Fila de encabezados con fondo rojo corporativo y texto blanco.
    - Por cada empleado: una fila de encabezado con su nombre (fondo salmon),
      sus registros de datos con bordes, y una fila de TOTAL con la suma
      de columnas numéricas (fondo rosa claro).
    - Fila en blanco entre empleados para facilitar la lectura.
    - Anchos de columna automáticos según el contenido.
    - Primera fila fija (freeze panes) para scroll largo.

    Parámetros
    ----------
    df : pd.DataFrame
        Datos a exportar. Debe contener la columna `group_col`.
    title : str
        Título de la hoja (se trunca a 30 caracteres).
    group_col : str, opcional
        Nombre de la columna de agrupación (default: "Nombre").

    Retorna
    -------
    bytes
        Contenido binario del archivo .xlsx listo para descargar.
    """
    wb = Workbook()
    ws = wb.active
    ws.title = title[:30]

    COLOR_HEADER = "8A2F1F"
    COLOR_GROUP  = "F0DDD6"
    COLOR_TOTAL  = "FFF3F0"

    thin = Side(style="thin", color="CCCCCC")
    brd  = Border(left=thin, right=thin, top=thin, bottom=thin)

    cols = list(df.columns)
    has_group = group_col in cols
    display_cols = [c for c in cols if c != group_col] if has_group else cols

    hdr_fill = PatternFill("solid", fgColor=COLOR_HEADER)
    hdr_font = Font(bold=True, color="FFFFFF", size=10)
    for ci, col in enumerate(display_cols, 1):
        cell = ws.cell(row=1, column=ci, value=col)
        cell.fill = hdr_fill
        cell.font = hdr_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = brd
    ws.row_dimensions[1].height = 18
    ws.freeze_panes = "A2"

    grp_fill = PatternFill("solid", fgColor=COLOR_GROUP)
    grp_font = Font(bold=True, color="3D1A0E", size=10)
    data_font = Font(size=9)
    tot_fill  = PatternFill("solid", fgColor=COLOR_TOTAL)
    tot_font  = Font(bold=True, size=9)

    row_num = 2
    groups = df.groupby(group_col, sort=False) if has_group else [(None, df)]
    for emp, grp in groups:
        if has_group:
            ws.merge_cells(start_row=row_num, start_column=1, end_row=row_num, end_column=len(display_cols))
            cell = ws.cell(row=row_num, column=1, value=str(emp))
            cell.fill = grp_fill
            cell.font = grp_font
            cell.alignment = Alignment(horizontal="left", vertical="center")
            ws.row_dimensions[row_num].height = 16
            row_num += 1

        num_cols_names = [c for c in display_cols if grp[c].dtype.kind in ("f", "i")]

        for _, data_row in grp.iterrows():
            for ci, col in enumerate(display_cols, 1):
                val = data_row[col]
                cell = ws.cell(row=row_num, column=ci, value=val)
                cell.font = data_font
                cell.border = brd
                if isinstance(val, float):
                    cell.number_format = "0.00"
            row_num += 1

        if has_group:
            for ci, col in enumerate(display_cols, 1):
                if col in num_cols_names:
                    val = grp[col].sum()
                    cell = ws.cell(row=row_num, column=ci, value=val)
                    cell.number_format = "0.00"
                else:
                    cell = ws.cell(row=row_num, column=ci, value="TOTAL" if ci == 1 else "")
                cell.fill = tot_fill
                cell.font = tot_font
                cell.border = brd
            row_num += 2

    for ci, col in enumerate(display_cols, 1):
        max_len = len(str(col))
        for row_cells in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=ci, max_col=ci):
            for cell in row_cells:
                if cell.value:
                    max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[get_column_letter(ci)].width = min(max_len + 2, 32)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()

def df_to_pdf(df, title):
    """
    Genera un archivo PDF simple con los datos de un DataFrame en tabla.

    Cada columna tiene ancho proporcional al número de columnas.
    Incluye una fila de título centrado y encabezados de columna con borde.

    Parámetros
    ----------
    df : pd.DataFrame
        Datos a exportar.
    title : str
        Título centrado en la parte superior del PDF.

    Retorna
    -------
    bytes
        Contenido binario del archivo .pdf listo para descargar.
    """
    pdf = FPDF()
    pdf.add_page()
    pdf.set_font("Arial", "B", 12)
    pdf.cell(0, 10, title, ln=1, align="C")
    pdf.set_font("Arial", size=8)
    col_width = pdf.w / (len(df.columns) + 1)
    row_height = pdf.font_size * 1.5
    for col in df.columns:
        pdf.cell(col_width, row_height, str(col), border=1)
    pdf.ln(row_height)
    for _, row in df.iterrows():
        for val in row:
            pdf.cell(col_width, row_height, str(val), border=1)
        pdf.ln(row_height)
    out = pdf.output(dest="S")
    if isinstance(out, str):
        return out.encode("latin1")
    if isinstance(out, bytearray):
        return bytes(out)
    return out

# ─────────────────────────────────────────────
# TABLAS RESUMEN
# ─────────────────────────────────────────────

# ── Tab: Por día ──
with tabs[0]:

    st.subheader("Resumen por persona / día")
    cols_show = ["Nombre", "Fecha", "Día", "Turno"] + [
        c for c in ["DO", "RNO", "HEDO", "HENO", "DOM", "RNF",
                    "HEDF", "HENF", "FEST", "RNDOM", "RDOM",
                    "ODOM", "ORNF", "OEDF", "OFEST", "TOTAL"]
        if c in dff.columns
    ]
    t_dia = dff[cols_show].copy()
    t_dia["Fecha"] = t_dia["Fecha"].dt.strftime("%d/%m/%Y")

    # Filtro extra por persona dentro de la tab
    per_dia = st.selectbox("Filtrar empleado (día)", ["Todos"] + sorted(dff["Nombre"].unique()), key="sel_dia")
    if per_dia != "Todos":
        t_dia = t_dia[t_dia["Nombre"] == per_dia]

    t_dia_sorted = t_dia.sort_values(["Nombre", "Fecha"]).reset_index(drop=True)
    _num_dia = [c for c in t_dia_sorted.columns if t_dia_sorted[c].dtype.kind in ("f", "i") and c != "Fecha"]
    _fmt_dia = {c: "{:.2f}" for c in _num_dia}
    # Un solo empleado: ocultar columna Nombre y mostrar como título
    if per_dia != "Todos":
        st.markdown(f"#### 👤 {per_dia}")
        st.dataframe(t_dia_sorted.drop(columns=["Nombre"]).style.format(_fmt_dia), use_container_width=True, height=400)
    else:
        st.dataframe(t_dia_sorted.style.format(_fmt_dia), use_container_width=True, height=400)

    # Exportar Excel estructurado
    xlsx_dia = df_to_excel_grouped(t_dia_sorted, "Por día")
    st.download_button(
        label="⬇ Descargar Excel",
        data=xlsx_dia,
        file_name="reporte_labor_dia.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )

# ── Tab: Por semana ──
with tabs[1]:

    st.subheader("Resumen por persona / semana")
    hora_cols_sum = [c for c in ["JORNADA", "DO", "RNO", "HEDO", "HENO", "DOM",
                                  "RNF", "HEDF", "HENF", "FEST", "RNDOM", "RDOM", "ODOM", "ORNF", "OFEST", "TOTAL"]
                     if c in dff.columns]
    t_sem = (
        dff.groupby(["Nombre", "Semana_etiqueta"])[hora_cols_sum]
        .sum()
        .reset_index()
        .rename(columns={"Semana_etiqueta": "Semana"})
    )
    per_sem = st.selectbox("Filtrar empleado (semana)", ["Todos"] + sorted(dff["Nombre"].unique()), key="sel_sem")
    if per_sem != "Todos":
        t_sem = t_sem[t_sem["Nombre"] == per_sem]

    t_sem_sorted = t_sem.sort_values(["Nombre", "Semana"]).reset_index(drop=True)
    num_cols = [c for c in t_sem_sorted.columns if c not in ("Nombre", "Semana")]
    if per_sem != "Todos":
        st.markdown(f"#### 👤 {per_sem}")
        df_sem_disp = t_sem_sorted.drop(columns=["Nombre"])
        st.dataframe(df_sem_disp.style.format({c: "{:.2f}" for c in num_cols}), use_container_width=True, height=400)
    else:
        st.dataframe(t_sem_sorted.style.format({c: "{:.2f}" for c in num_cols}), use_container_width=True, height=400)

    # Exportar Excel estructurado
    xlsx_sem = df_to_excel_grouped(t_sem_sorted, "Por semana")
    st.download_button(
        label="⬇ Descargar Excel",
        data=xlsx_sem,
        file_name="reporte_labor_semana.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )

# ── Tab: Por mes ──
with tabs[2]:

    st.subheader("Resumen por persona / mes")
    t_mes = (
        dff.groupby(["Nombre", "Mes"])[hora_cols_sum]
        .sum()
        .reset_index()
    )
    per_mes = st.selectbox("Filtrar empleado (mes)", ["Todos"] + sorted(dff["Nombre"].unique()), key="sel_mes")
    if per_mes != "Todos":
        t_mes = t_mes[t_mes["Nombre"] == per_mes]

    t_mes_sorted = t_mes.sort_values("Nombre").reset_index(drop=True)
    num_cols_mes = [c for c in t_mes_sorted.columns if c not in ("Nombre", "Mes")]
    if per_mes != "Todos":
        st.markdown(f"#### 👤 {per_mes}")
        df_mes_disp = t_mes_sorted.drop(columns=["Nombre"])
        st.dataframe(df_mes_disp.style.format({c: "{:.2f}" for c in num_cols_mes}), use_container_width=True, height=400)
    else:
        st.dataframe(t_mes_sorted.style.format({c: "{:.2f}" for c in num_cols_mes}), use_container_width=True, height=400)

    # Exportar Excel estructurado
    xlsx_mes = df_to_excel_grouped(t_mes_sorted, "Por mes")
    st.download_button(
        label="⬇ Descargar Excel",
        data=xlsx_mes,
        file_name="reporte_labor_mes.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )

# ── Tab: Detalle completo ──
with tabs[3]:

    st.subheader("Datos completos")
    det = dff.copy()
    det["Fecha"] = det["Fecha"].dt.strftime("%d/%m/%Y")
    per_det = st.selectbox("Filtrar empleado (detalle)", ["Todos"] + sorted(dff["Nombre"].unique()), key="sel_det")
    if per_det != "Todos":
        det = det[det["Nombre"] == per_det]

    det_sorted = det.sort_values(["Nombre", "Fecha"]).reset_index(drop=True)
    _num_det = [c for c in det_sorted.columns if det_sorted[c].dtype.kind in ("f", "i") and c != "Fecha"]
    _fmt_det = {c: "{:.2f}" for c in _num_det}
    if per_det != "Todos":
        st.markdown(f"#### 👤 {per_det}")
        st.dataframe(det_sorted.drop(columns=["Nombre"]).style.format(_fmt_det), use_container_width=True, height=500)
    else:
        st.dataframe(det_sorted.style.format(_fmt_det), use_container_width=True, height=500)

    # Botón de descarga Excel
    xlsx_det = df_to_excel_grouped(det_sorted, "Detalle completo")
    st.download_button(
        label="⬇ Descargar Excel",
        data=xlsx_det,
        file_name="reporte_labor_detalle.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )

# ── Tab: Cumplimiento Ley 2466 ──
with tabs[4]:
    st.subheader("⚖️ Cumplimiento normativo – Horas extras")
    st.markdown(
        """
        **Norma aplicada:** Ley 2466 de 2025 / Código Sustantivo del Trabajo
        - 🔴 Máximo **2 horas extras** por día por empleado.
        - 🔴 Máximo **12 horas extras** por semana por empleado.

        Las filas resaltadas en 🚨 indican incumplimiento del límite legal.
        """,
    )

    cump_diario, cump_semanal = calcular_cumplimiento(dff)
    # Redondear columnas numéricas del reporte de cumplimiento
    for _col in ["H. Extra del día", "Exceso diario"]:
        if _col in cump_diario.columns:
            cump_diario[_col] = cump_diario[_col].round(2)
    for _col in ["H. Extra semana", "Exceso semanal"]:
        if _col in cump_semanal.columns:
            cump_semanal[_col] = cump_semanal[_col].round(2)

    # Guardar versiones sin filtro de empleado para análisis por tienda
    _cump_dia_all = cump_diario.copy()
    _cump_sem_all = cump_semanal.copy()

    # Mapear Nombre → Tienda desde dff
    _nombre_grupo = dff[["Nombre", "Grupo"]].drop_duplicates().set_index("Nombre")["Grupo"].to_dict()
    _cump_dia_all["Tienda"] = _cump_dia_all["Nombre"].map(_nombre_grupo).fillna("(Sin grupo)")
    _cump_sem_all["Tienda"] = _cump_sem_all["Nombre"].map(_nombre_grupo).fillna("(Sin grupo)")

    # ── Filtro por empleado ──
    per_cump = st.selectbox(
        "Filtrar empleado",
        ["Todos"] + sorted(dff["Nombre"].unique()),
        key="sel_cump",
    )
    if per_cump != "Todos":
        cump_diario  = cump_diario[cump_diario["Nombre"] == per_cump]
        cump_semanal = cump_semanal[cump_semanal["Nombre"] == per_cump]

    # ── Resumen de alertas ──
    n_excesos_dia = (cump_diario["Estado"] == "🚨 EXCEDE LÍMITE").sum()
    n_excesos_sem = (cump_semanal["Estado"] == "🚨 EXCEDE LÍMITE").sum()
    c1, c2 = st.columns(2)
    with c1:
        if n_excesos_dia > 0:
            st.error(f"🚨 {n_excesos_dia} día(s) con más de 2 h extras")
        else:
            st.success("✅ Sin infracciones diarias")
    with c2:
        if n_excesos_sem > 0:
            st.error(f"🚨 {n_excesos_sem} semana(s) con más de 12 h extras")
        else:
            st.success("✅ Sin infracciones semanales")

    # ── Filtro de vista: solo incumplimientos o todos ──
    solo_incumplimientos = st.toggle(
        "Mostrar solo incumplimientos",
        value=True,
        help="Activa para ver únicamente los registros que exceden el límite legal.",
    )

    # Colores condicionales: rojo si excede, verde si OK
    def _color_estado(val):
        if "EXCEDE" in str(val):
            return "background-color: #FFEBEE; color: #B71C1C; font-weight: bold"
        return "background-color: #E8F5E9; color: #1B5E20"

    # ── Tabla diaria ──
    st.markdown("#### Detalle diario")
    disp_dia = cump_diario.copy()
    if solo_incumplimientos:
        disp_dia = disp_dia[disp_dia["Estado"] == "🚨 EXCEDE LÍMITE"]
    if per_cump != "Todos":
        disp_dia = disp_dia.drop(columns=["Nombre"])
    if disp_dia.empty:
        st.success("✅ Sin infracciones diarias en el período seleccionado.")
    else:
        _num_dia_c = [c for c in disp_dia.columns if disp_dia[c].dtype.kind in ("f", "i")]
        st.dataframe(
            disp_dia.style
                .format({c: "{:.2f}" for c in _num_dia_c})
                .map(_color_estado, subset=["Estado"]),
            use_container_width=True,
            height=min(60 + len(disp_dia) * 36, 420),
        )
        _xlsx_cump_dia = df_to_excel_grouped(disp_dia.reset_index(drop=True), "Cumplimiento diario", group_col="_none_")
        st.download_button("⬇ Descargar Excel (diario)", data=_xlsx_cump_dia,
            file_name="cumplimiento_diario.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            key="dl_cump_dia")

    # ── Tabla semanal ──
    st.markdown("#### Detalle semanal")
    disp_sem = cump_semanal.copy()
    if solo_incumplimientos:
        disp_sem = disp_sem[disp_sem["Estado"] == "🚨 EXCEDE LÍMITE"]
    if per_cump != "Todos":
        disp_sem = disp_sem.drop(columns=["Nombre"])
    if disp_sem.empty:
        st.success("✅ Sin infracciones semanales en el período seleccionado.")
    else:
        _num_sem_c = [c for c in disp_sem.columns if disp_sem[c].dtype.kind in ("f", "i")]
        st.dataframe(
            disp_sem.style
                .format({c: "{:.2f}" for c in _num_sem_c})
                .map(_color_estado, subset=["Estado"]),
            use_container_width=True,
            height=min(60 + len(disp_sem) * 36, 360),
        )
        _xlsx_cump_sem = df_to_excel_grouped(disp_sem.reset_index(drop=True), "Cumplimiento semanal", group_col="_none_")
        st.download_button("⬇ Descargar Excel (semanal)", data=_xlsx_cump_sem,
            file_name="cumplimiento_semanal.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            key="dl_cump_sem")

    # ── Descarga Excel con las dos hojas (siempre exporta todo, no filtrado) ──
    xlsx_cump = excel_cumplimiento(cump_diario, cump_semanal)
    st.download_button(
        label="⬇ Descargar reporte de cumplimiento (Excel)",
        data=xlsx_cump,
        file_name="cumplimiento_horas_extras.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )

    # ── Análisis de incumplimiento por Tienda ──
    st.divider()
    st.markdown("### 🏪 Incumplimiento por Tienda")
    st.caption("Consolidado de infracciones de horas extras agrupado por tienda/grupo.")

    _exc_dia_tienda = (
        _cump_dia_all[_cump_dia_all["Estado"] == "🚨 EXCEDE LÍMITE"]
        .groupby("Tienda")
        .agg(
            Infracciones_diarias=("Nombre", "count"),
            Empleados_con_exceso=("Nombre", "nunique"),
            Exceso_total_horas=("Exceso diario", "sum"),
        )
        .reset_index()
        .sort_values("Infracciones_diarias", ascending=False)
        .reset_index(drop=True)
    )
    _exc_dia_tienda["Exceso_total_horas"] = _exc_dia_tienda["Exceso_total_horas"].round(2)

    _exc_sem_tienda = (
        _cump_sem_all[_cump_sem_all["Estado"] == "🚨 EXCEDE LÍMITE"]
        .groupby("Tienda")
        .agg(
            Infracciones_semanales=("Nombre", "count"),
            Empleados_con_exceso=("Nombre", "nunique"),
            Exceso_total_horas=("Exceso semanal", "sum"),
        )
        .reset_index()
        .sort_values("Infracciones_semanales", ascending=False)
        .reset_index(drop=True)
    )
    _exc_sem_tienda["Exceso_total_horas"] = _exc_sem_tienda["Exceso_total_horas"].round(2)

    # KPIs por tienda
    _t_inf1, _t_inf2 = st.columns(2)
    with _t_inf1:
        _n_tiendas_dia = len(_exc_dia_tienda)
        if _n_tiendas_dia > 0:
            st.error(f"🚨 {_n_tiendas_dia} tienda(s) con infracciones diarias")
        else:
            st.success("✅ Ninguna tienda supera el límite diario")
    with _t_inf2:
        _n_tiendas_sem = len(_exc_sem_tienda)
        if _n_tiendas_sem > 0:
            st.error(f"🚨 {_n_tiendas_sem} tienda(s) con infracciones semanales")
        else:
            st.success("✅ Ninguna tienda supera el límite semanal")

    _sub_tienda = st.tabs(["Infracciones diarias por tienda", "Infracciones semanales por tienda"])

    # ── Subtab: Diario por tienda ──
    with _sub_tienda[0]:
        if _exc_dia_tienda.empty:
            st.success("✅ Ninguna tienda con infracciones diarias.")
        else:
            # Gráfica de barras horizontal
            _fig_h_dia = max(300, len(_exc_dia_tienda) * 48 + 80)
            _fig_tid = px.bar(
                _exc_dia_tienda.sort_values("Infracciones_diarias"),
                x="Infracciones_diarias",
                y="Tienda",
                orientation="h",
                color="Infracciones_diarias",
                color_continuous_scale=[[0, "#FFEBEE"], [0.5, "#EF9A9A"], [1, "#B71C1C"]],
                text="Infracciones_diarias",
                height=_fig_h_dia,
                labels={"Infracciones_diarias": "N° infracciones", "Tienda": "Tienda"},
                title="Tiendas con infracciones diarias (> 2 h extras/día)",
            )
            _fig_tid.update_traces(textposition="outside")
            _fig_tid.update_layout(
                coloraxis_showscale=False,
                plot_bgcolor="rgba(0,0,0,0)",
                paper_bgcolor="rgba(0,0,0,0)",
                font_size=12,
                margin=dict(l=10, r=60, t=50, b=30),
            )
            st.plotly_chart(_fig_tid, use_container_width=True)

            st.markdown("##### Detalle por tienda")
            _exc_dia_tienda.insert(0, "#", range(1, len(_exc_dia_tienda) + 1))
            _exc_dia_tienda_disp = _exc_dia_tienda.rename(columns={
                "Infracciones_diarias": "N° infracciones",
                "Empleados_con_exceso": "Empleados afectados",
                "Exceso_total_horas": "Total horas exceso",
            })
            def _color_tienda(row):
                return ["background-color:#FFEBEE; color:#B71C1C; font-weight:bold"] * len(row)
            st.dataframe(
                _exc_dia_tienda_disp.style
                    .apply(_color_tienda, axis=1)
                    .format({"Total horas exceso": "{:.2f}"}),
                use_container_width=True,
                hide_index=True,
                height=min(80 + len(_exc_dia_tienda_disp) * 38, 420),
            )
            _xlsx_tienda_dia = df_to_excel_grouped(
                _exc_dia_tienda_disp.drop(columns=["#"]), "Infracciones diarias por tienda", group_col="_none_"
            )
            st.download_button(
                "⬇ Descargar Excel (infracciones diarias por tienda)",
                data=_xlsx_tienda_dia,
                file_name="infracciones_diarias_tienda.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                key="dl_inf_dia_tienda",
            )

            # Detalle de empleados infractores de la tienda seleccionada
            st.markdown("##### Empleados con infracciones diarias por tienda")
            _tiendas_inf_dia = ["Todas"] + sorted(_cump_dia_all[_cump_dia_all["Estado"] == "🚨 EXCEDE LÍMITE"]["Tienda"].unique())
            _sel_t_dia = st.selectbox("Ver tienda", _tiendas_inf_dia, key="sel_tienda_inf_dia")
            _det_emp_dia = _cump_dia_all[_cump_dia_all["Estado"] == "🚨 EXCEDE LÍMITE"].copy()
            if _sel_t_dia != "Todas":
                _det_emp_dia = _det_emp_dia[_det_emp_dia["Tienda"] == _sel_t_dia]
            _det_emp_dia = _det_emp_dia.sort_values(["Tienda", "Nombre", "Fecha"]).reset_index(drop=True)
            _num_det_dia = [c for c in _det_emp_dia.columns if _det_emp_dia[c].dtype.kind in ("f", "i")]
            st.dataframe(
                _det_emp_dia.style
                    .format({c: "{:.2f}" for c in _num_det_dia})
                    .map(_color_estado, subset=["Estado"]),
                use_container_width=True,
                height=min(80 + len(_det_emp_dia) * 36, 420),
            )

    # ── Subtab: Semanal por tienda ──
    with _sub_tienda[1]:
        if _exc_sem_tienda.empty:
            st.success("✅ Ninguna tienda con infracciones semanales.")
        else:
            _fig_h_sem = max(300, len(_exc_sem_tienda) * 48 + 80)
            _fig_tis = px.bar(
                _exc_sem_tienda.sort_values("Infracciones_semanales"),
                x="Infracciones_semanales",
                y="Tienda",
                orientation="h",
                color="Infracciones_semanales",
                color_continuous_scale=[[0, "#FFEBEE"], [0.5, "#EF9A9A"], [1, "#B71C1C"]],
                text="Infracciones_semanales",
                height=_fig_h_sem,
                labels={"Infracciones_semanales": "N° infracciones", "Tienda": "Tienda"},
                title="Tiendas con infracciones semanales (> 12 h extras/semana)",
            )
            _fig_tis.update_traces(textposition="outside")
            _fig_tis.update_layout(
                coloraxis_showscale=False,
                plot_bgcolor="rgba(0,0,0,0)",
                paper_bgcolor="rgba(0,0,0,0)",
                font_size=12,
                margin=dict(l=10, r=60, t=50, b=30),
            )
            st.plotly_chart(_fig_tis, use_container_width=True)

            st.markdown("##### Detalle por tienda")
            _exc_sem_tienda.insert(0, "#", range(1, len(_exc_sem_tienda) + 1))
            _exc_sem_tienda_disp = _exc_sem_tienda.rename(columns={
                "Infracciones_semanales": "N° infracciones",
                "Empleados_con_exceso": "Empleados afectados",
                "Exceso_total_horas": "Total horas exceso",
            })
            st.dataframe(
                _exc_sem_tienda_disp.style
                    .apply(_color_tienda, axis=1)
                    .format({"Total horas exceso": "{:.2f}"}),
                use_container_width=True,
                hide_index=True,
                height=min(80 + len(_exc_sem_tienda_disp) * 38, 420),
            )
            _xlsx_tienda_sem = df_to_excel_grouped(
                _exc_sem_tienda_disp.drop(columns=["#"]), "Infracciones semanales por tienda", group_col="_none_"
            )
            st.download_button(
                "⬇ Descargar Excel (infracciones semanales por tienda)",
                data=_xlsx_tienda_sem,
                file_name="infracciones_semanales_tienda.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                key="dl_inf_sem_tienda",
            )

            # Detalle de empleados infractores de la tienda seleccionada
            st.markdown("##### Empleados con infracciones semanales por tienda")
            _tiendas_inf_sem = ["Todas"] + sorted(_cump_sem_all[_cump_sem_all["Estado"] == "🚨 EXCEDE LÍMITE"]["Tienda"].unique())
            _sel_t_sem = st.selectbox("Ver tienda", _tiendas_inf_sem, key="sel_tienda_inf_sem")
            _det_emp_sem = _cump_sem_all[_cump_sem_all["Estado"] == "🚨 EXCEDE LÍMITE"].copy()
            if _sel_t_sem != "Todas":
                _det_emp_sem = _det_emp_sem[_det_emp_sem["Tienda"] == _sel_t_sem]
            _det_emp_sem = _det_emp_sem.sort_values(["Tienda", "Nombre"]).reset_index(drop=True)
            _num_det_sem = [c for c in _det_emp_sem.columns if _det_emp_sem[c].dtype.kind in ("f", "i")]
            st.dataframe(
                _det_emp_sem.style
                    .format({c: "{:.2f}" for c in _num_det_sem})
                    .map(_color_estado, subset=["Estado"]),
                use_container_width=True,
                height=min(80 + len(_det_emp_sem) * 36, 420),
            )

# ── Tab: Total laborado ──
with tabs[5]:
    st.subheader("Total de tiempo laborado")

    # JORNADA se excluye porque es columna de referencia del turno, no componente del TOTAL.
    # El TOTAL de Siesa = DO + recargos + extras + dom + festivos (sin JORNADA).
    _all_hour_cols_t = [c for c in HORA_COLS.keys() if c not in ("TOTAL", "JORNADA") and c in dff.columns]
    _extras_ord_cols  = [c for c in ["HEDO", "HENO"] if c in dff.columns]
    _extras_fest_cols = [c for c in ["HEDF", "HENF"] if c in dff.columns]
    _dom_fest_cols    = [c for c in ["DOM", "FEST", "RNDOM", "RDOM"] if c in dff.columns]
    _recargo_cols     = [c for c in ["RNO", "RNF", "ODOM", "ORNF", "OEDF", "OFEST"] if c in dff.columns]

    _grp_t = dff.groupby("Nombre")[_all_hour_cols_t + (["TOTAL"] if "TOTAL" in dff.columns else [])].sum().reset_index()

    _grp_t["DO"]               = _grp_t["DO"].round(2) if "DO" in _grp_t.columns else 0
    _grp_t["Extras_Ordinarias"] = _grp_t[[c for c in _extras_ord_cols if c in _grp_t.columns]].sum(axis=1).round(2)
    _grp_t["Extras_Festivas"]   = _grp_t[[c for c in _extras_fest_cols if c in _grp_t.columns]].sum(axis=1).round(2)
    _grp_t["Dom_Festivo"]       = _grp_t[[c for c in _dom_fest_cols if c in _grp_t.columns]].sum(axis=1).round(2)
    _grp_t["Recargos"]          = _grp_t[[c for c in _recargo_cols if c in _grp_t.columns]].sum(axis=1).round(2)
    if "TOTAL" in _grp_t.columns:
        _grp_t["TOTAL"] = _grp_t["TOTAL"].round(2)

    _fmt_t = {c: "{:.2f}" for c in _grp_t.columns if _grp_t[c].dtype.kind in ("f", "i")}

    subt = st.tabs(["Resumen", "Ordinario vs Extra", "Detalle completo"])

    # ── Subtab: Resumen agrupado ──
    with subt[0]:
        # Columnas individuales — incluye también ODOM/ORNF/OEDF/OFEST (algunos archivos las tienen)
        _he_cols_exist = [c for c in ["DO", "RNO", "HEDO", "HENO", "DOM", "FEST",
                                      "RNF", "HEDF", "HENF", "RNDOM", "RDOM",
                                      "ODOM", "ORNF", "OEDF", "OFEST"] if c in _grp_t.columns]
        _summary_cols = ["Nombre"] + _he_cols_exist
        if "TOTAL" in _grp_t.columns:
            _summary_cols.append("TOTAL")
        _disp_res = _grp_t[_summary_cols].sort_values("Nombre").reset_index(drop=True)
        _disp_res = _disp_res.rename(columns={
            "DO":    "DO (Jornada ord.)",
            "RNO":   "RNO (Rec. noct. ord.)",
            "HEDO":  "HEDO (H.E. Diurna Ord.)",
            "HENO":  "HENO (H.E. Noct. Ord.)",
            "DOM":   "DOM (Dominical)",
            "FEST":  "FEST (Festivo)",
            "RNF":   "RNF (Rec. noct. fest.)",
            "HEDF":  "HEDF (H.E. Diurna Fest.)",
            "HENF":  "HENF (H.E. Noct. Fest.)",
            "RNDOM": "RNDOM (Rec. noct. dom.)",
            "RDOM":  "RDOM (Rec. dom.)",
            "ODOM":  "ODOM (H.E. dom.)",
            "ORNF":  "ORNF (Rec. noct. fest.)",
            "OEDF":  "OEDF (H.E. diurna fest.)",
            "OFEST": "OFEST (Otras fest.)",
            "TOTAL": "TOTAL",
        })
        st.caption("Cada columna es la suma de todas las horas del período. TOTAL = suma de todas las categorías.")
        st.dataframe(_disp_res.style.format({c: "{:.2f}" for c in _disp_res.columns if _disp_res[c].dtype.kind in ("f","i")}),
                     use_container_width=True, height=min(60 + len(_disp_res) * 36, 500))
        _xlsx_res = df_to_excel_grouped(_disp_res, "Resumen laborado")
        st.download_button("⬇ Descargar Excel", data=_xlsx_res, file_name="total_resumen.xlsx",
                           mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", key="dl_tot_res")

    # ── Subtab: Ordinario vs Extra ──
    with subt[1]:
        _norm_cols = [c for c in _all_hour_cols_t if c not in (_extras_ord_cols + _extras_fest_cols)]
        _grp_t["Tiempo_Ordinario"] = _grp_t[[c for c in _norm_cols if c in _grp_t.columns]].sum(axis=1).round(2)
        _grp_t["Tiempo_Extra"]     = (_grp_t["Extras_Ordinarias"] + _grp_t["Extras_Festivas"]).round(2)
        _disp_ord = _grp_t[["Nombre", "Tiempo_Ordinario", "Tiempo_Extra"] + (["TOTAL"] if "TOTAL" in _grp_t.columns else [])].sort_values("Nombre").reset_index(drop=True)
        st.dataframe(_disp_ord.style.format({c: "{:.2f}" for c in _disp_ord.columns if _disp_ord[c].dtype.kind in ("f","i")}),
                     use_container_width=True, height=min(60 + len(_disp_ord) * 36, 500))
        _xlsx_ord = df_to_excel_grouped(_disp_ord, "Total Ordinario vs Extra")
        st.download_button("⬇ Descargar Excel", data=_xlsx_ord, file_name="total_ordinario_extra.xlsx",
                           mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", key="dl_tot_ord")

    # ── Subtab: Detalle completo ──
    with subt[2]:
        _det_cols = ["Nombre"] + [c for c in _all_hour_cols_t if c in _grp_t.columns] + (["TOTAL"] if "TOTAL" in _grp_t.columns else [])
        _disp_ext = _grp_t[_det_cols].sort_values("Nombre").reset_index(drop=True)
        st.caption("Todas las columnas de horas individuales del reporte Siesa.")
        st.dataframe(_disp_ext.style.format({c: "{:.2f}" for c in _disp_ext.columns if _disp_ext[c].dtype.kind in ("f","i")}),
                     use_container_width=True, height=min(60 + len(_disp_ext) * 36, 500))
        _xlsx_ext = df_to_excel_grouped(_disp_ext, "Detalle completo laborado")
        st.download_button("⬇ Descargar Excel", data=_xlsx_ext, file_name="total_detalle_completo.xlsx",
                           mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", key="dl_tot_ext")

# ─────────────────────────────────────────────
# GRÁFICA COMPARATIVA – Distribución de tipos de horas
# ─────────────────────────────────────────────
st.divider()
st.subheader("Distribución de tipos de horas por empleado")

dist_cols = [c for c in ["JORNADA", "HEDO", "HENO", "DOM", "HEDF", "HENF", "FEST"] if c in dff.columns]

# --- Definir altura de la gráfica antes de usarla ---
fig2_height = 420
if is_mobile:
    fig2_height = 260

dist_df = dff.groupby("Nombre")[dist_cols].sum().reset_index()
dist_melt = dist_df.melt(id_vars="Nombre", var_name="Tipo", value_name="Horas")
dist_melt["Tipo_desc"] = dist_melt["Tipo"].map(lambda x: HORA_COLS.get(x, x))
dist_melt_filtered = dist_melt[dist_melt["Horas"] > 0]

# --- Lógica para la segunda gráfica ---
if is_mobile:
    # Mostrar solo top N empleados en móvil, siempre como dona/pie
    if dist_melt_filtered.shape[0] > 0 and 'Nombre' in dist_melt_filtered.columns:
        top_empleados2 = dist_melt_filtered.groupby('Nombre')['Horas'].sum().nlargest(TOP_N).index
        dist_melt_filtered_mobile = dist_melt_filtered[dist_melt_filtered['Nombre'].isin(top_empleados2)]
        fig2 = px.pie(
            dist_melt_filtered_mobile,
            values="Horas",
            names="Nombre",
            hole=0.45,
            height=fig2_height,
            title=f"Distribución de horas por empleado (top {TOP_N})",
        )
        fig2.update_traces(
            textinfo='percent',
            textposition='inside',
            insidetextorientation='radial',
        )
        fig2.update_layout(
            legend_title="Empleado",
            font_size=12,
            showlegend=True,
            legend=dict(orientation='v', x=1.02, y=0.5),
        )
        st.info(f"Mostrando solo los {TOP_N} empleados con más horas para mejor visualización en móvil.")
    else:
        fig2 = px.pie()  # Gráfica vacía si no hay datos
else:
    # Desktop: dona si solo hay un empleado, barras apiladas si hay varios
    if dist_df.shape[0] == 1:
        fig2 = px.pie(
            dist_melt_filtered,
            values="Horas",
            names="Tipo_desc",
            hole=0.4,
            labels={"Horas": "Horas", "Tipo_desc": "Tipo"},
            height=fig2_height,
            title="Composición de horas por empleado (período seleccionado)",
        )
        fig2.update_layout(
            plot_bgcolor="rgba(0,0,0,0)",
            paper_bgcolor="rgba(0,0,0,0)",
            legend_title="Tipo de hora",
        )
    else:
        fig2 = px.bar(
            dist_melt_filtered,
            x="Nombre",
            y="Horas",
            color="Tipo_desc",
            barmode="stack",
            height=fig2_height,
            labels={"Nombre": "Empleado", "Horas": "Horas", "Tipo_desc": "Tipo"},
            title="Composición de horas por empleado (período seleccionado)",
        )
        fig2.update_layout(
            xaxis_tickangle=-30,
            plot_bgcolor="rgba(0,0,0,0)",
            paper_bgcolor="rgba(0,0,0,0)",
            legend_title="Tipo de hora",
            font_size=12,
        )

fig2.update_layout(height=fig2_height)
st.plotly_chart(fig2, use_container_width=True)

# ── Tab: SQL Histórico (aditivo) ──
with tabs[7]:
    st.subheader("🔌 SQL Histórico y Validación")
    st.caption("Esta sección agrega consulta histórica SQL sin reemplazar el flujo actual de Excel.")

    if pytds is None:
        st.error("Falta dependencia de SQL Server. Instala `python-tds` en el entorno de Contabilidad.")
    else:
        cfg = _db_cfg()
        missing = _db_missing(cfg)
        if missing:
            st.warning("Faltan variables de entorno/secrets para conexión SQL: " + ", ".join(missing))
            st.info("Configúralas en `.streamlit/secrets.toml` o variables de entorno. La conexión es automática.")
        else:
            try:
                ok = _sql_test_connection(cfg)
                st.success("Conexión SQL automática activa." if ok else "Conexión SQL no disponible.")
            except Exception as e:
                st.error(f"No se pudo validar conexión SQL: {e}")
                ok = False

            if ok:
                st.markdown("#### Consulta SQL por mes completo")
                _month_range = pd.period_range(end=pd.Timestamp.today().to_period("M"), periods=72, freq="M")
                _month_opts = [str(m) for m in _month_range]
                _default_ini_idx = max(0, len(_month_opts) - 6)

                col_sql_mes_1, col_sql_mes_2 = st.columns(2)
                with col_sql_mes_1:
                    sql_mes_ini = st.selectbox(
                        "Mes inicial",
                        options=_month_opts,
                        index=_default_ini_idx,
                        key="sql_mes_ini",
                    )
                with col_sql_mes_2:
                    sql_mes_fin = st.selectbox(
                        "Mes final",
                        options=_month_opts,
                        index=len(_month_opts) - 1,
                        key="sql_mes_fin",
                    )

                csql1, csql2 = st.columns(2)
                with csql1:
                    if st.button("Consultar SQL histórico", use_container_width=True, key="btn_sql_hist"):
                        _p_ini = pd.Period(sql_mes_ini, freq="M")
                        _p_fin = pd.Period(sql_mes_fin, freq="M")
                        if _p_ini > _p_fin:
                            st.warning("El mes inicial no puede ser mayor al mes final.")
                        else:
                            try:
                                sql_df = _sql_fetch_hours(cfg, _p_ini.start_time.date(), _p_fin.end_time.date())
                                st.session_state["sql_hist_df"] = sql_df
                                st.success(f"Consulta SQL completada: {len(sql_df):,} filas.")
                            except Exception as e:
                                st.error(f"Error consultando SQL: {e}")
                with csql2:
                    if st.button("Limpiar consulta SQL", use_container_width=True, key="btn_sql_clear"):
                        st.session_state.pop("sql_hist_df", None)
                        st.rerun()

                sql_hist_df = st.session_state.get("sql_hist_df", pd.DataFrame())
                if sql_hist_df.empty:
                    st.info("Ejecuta 'Consultar SQL histórico' para traer horas desde base de datos.")
                else:
                    st.markdown("#### Alineación obligatoria: personal de tienda")
                    _sql_before = len(sql_hist_df)
                    _docs_archivo = set()
                    if "Documento" in dff.columns and "Documento" in sql_hist_df.columns:
                        # Base esperada de personal tienda: filtros de tienda/empleado del archivo,
                        # sin restringir por fecha para no perder empleados de nómina del mes.
                        _df_tienda_base = df[
                            df["Grupo"].fillna("(Sin grupo)").isin(sel_grupos)
                            & df["Nombre"].isin(sel_personas)
                        ].copy()
                        _docs_archivo = set(_normalize_doc_series(_df_tienda_base["Documento"]).dropna().tolist())
                        sql_hist_df["Documento"] = _normalize_doc_series(sql_hist_df["Documento"])
                        sql_hist_df = sql_hist_df[sql_hist_df["Documento"].isin(_docs_archivo)].copy()
                        st.caption(
                            f"SQL filtrado solo a personal de tienda seleccionado en el archivo: {_sql_before:,} → {len(sql_hist_df):,} filas."
                        )

                    # Auditoría base de documentos/fila para encontrar brechas
                    _docs_sql = set(sql_hist_df["Documento"].dropna().astype(str).tolist()) if "Documento" in sql_hist_df.columns else set()
                    _faltan_sql = sorted(list(_docs_archivo - _docs_sql))
                    _extras_sql = sorted(list(_docs_sql - _docs_archivo))
                    _a1, _a2, _a3, _a4 = st.columns(4)
                    _a1.metric("Filas SQL (post filtro tienda)", f"{len(sql_hist_df):,}")
                    _a2.metric("Docs archivo (tienda)", f"{len(_docs_archivo):,}")
                    _a3.metric("Docs SQL", f"{len(_docs_sql):,}")
                    _a4.metric("Docs faltantes SQL", f"{len(_faltan_sql):,}")
                    st.caption(f"Control de plantilla de tienda esperada: {len(_docs_archivo):,} empleados/documentos.")
                    with st.expander("Ver auditoría de documentos (faltantes/extras)"):
                        _left, _right = st.columns(2)
                        with _left:
                            st.markdown("**Documentos del archivo que NO aparecen en SQL**")
                            st.dataframe(pd.DataFrame({"Documento": _faltan_sql}), use_container_width=True, hide_index=True)
                        with _right:
                            st.markdown("**Documentos en SQL que NO están en archivo**")
                            st.dataframe(pd.DataFrame({"Documento": _extras_sql}), use_container_width=True, hide_index=True)

                    _cmp_cols = ["DO", "RNO", "HEDO", "HENO", "DOM", "RNF", "HEDF", "HENF", "FEST", "RNDOM", "RDOM", "ODOM", "ORNF", "OFEST", "TOTAL"]
                    for _c in _cmp_cols:
                        if _c not in sql_hist_df.columns:
                            sql_hist_df[_c] = 0.0

                    _sql_mes_opts = sorted(sql_hist_df["Mes"].dropna().unique().tolist())
                    st.markdown("#### Resumen mensual SQL")
                    st.info(
                        "**DO en SQL ≠ DO en Excel.** El concepto SALBAS en nómina registra las "
                        "**horas contratadas** (días hábiles × turno), no las horas físicamente "
                        "marcadas en el reloj. Además, las horas ordinarias del domingo (DOM del "
                        "Excel) van incluidas dentro de SALBAS porque en nómina no existe un "
                        "concepto separado para ellas. El resto de tipos (HEDO, HENO, RNO, RNF, "
                        "FEST, HENF, etc.) sí tienen conceptos directos y deben coincidir "
                        "aproximadamente con el Excel.",
                        icon="ℹ️",
                    )
                    sql_month = sql_hist_df.groupby("Mes", as_index=False)[_cmp_cols].sum()
                    st.dataframe(
                        sql_month.sort_values("Mes").style.format({c: "{:.2f}" for c in sql_month.columns if c != "Mes"}),
                        use_container_width=True,
                        hide_index=True,
                    )

                    # ── Reconciliación SQL vs Excel por mes ──────────────────────
                    st.markdown("#### Reconciliación SQL vs Excel (por mes del archivo)")
                    _excl_recon_cols = ["DO", "RNO", "HEDO", "HENO", "DOM", "RNF",
                                        "HEDF", "HENF", "FEST", "RNDOM", "RDOM",
                                        "ODOM", "ORNF", "OFEST", "TOTAL"]
                    _dff_mes_norm = dff.copy()
                    _dff_mes_norm["Documento"] = _normalize_doc_series(_dff_mes_norm["Documento"])
                    recon_rows = []
                    for _mes_r in sorted(sql_hist_df["Mes"].dropna().unique()):
                        _sql_r = sql_hist_df[sql_hist_df["Mes"] == _mes_r]
                        _doc_r = _dff_mes_norm[_dff_mes_norm["Mes"] == _mes_r] if "Mes" in _dff_mes_norm.columns else pd.DataFrame()
                        row = {"Mes": _mes_r}
                        for _c in _excl_recon_cols:
                            _sql_val = float(_sql_r[_c].sum()) if _c in _sql_r.columns else 0.0
                            _doc_val = float(_doc_r[_c].sum()) if (not _doc_r.empty and _c in _doc_r.columns) else 0.0
                            row[f"SQL_{_c}"] = round(_sql_val, 2)
                            row[f"XLS_{_c}"] = round(_doc_val, 2)
                            row[f"Δ_{_c}"] = round(_sql_val - _doc_val, 2)
                        # DO ajustado: SQL DO vs Excel (DO + DOM combinados)
                        _sql_do = float(_sql_r["DO"].sum()) if "DO" in _sql_r.columns else 0.0
                        _doc_do_dom = (
                            (float(_doc_r["DO"].sum()) if "DO" in _doc_r.columns else 0.0) +
                            (float(_doc_r["DOM"].sum()) if "DOM" in _doc_r.columns else 0.0)
                        ) if not _doc_r.empty else 0.0
                        row["SQL_DO"] = round(_sql_do, 2)
                        row["XLS_DO+DOM"] = round(_doc_do_dom, 2)
                        row["Δ_DO_ajustado"] = round(_sql_do - _doc_do_dom, 2)
                        recon_rows.append(row)

                    if recon_rows:
                        _recon_df = pd.DataFrame(recon_rows)
                        # Vista compacta: mostrar solo columnas con diferencias significativas
                        _view_cols = ["Mes", "SQL_DO", "XLS_DO+DOM", "Δ_DO_ajustado",
                                      "SQL_DOM", "XLS_DOM", "Δ_DOM",
                                      "SQL_HEDO", "XLS_HEDO", "Δ_HEDO",
                                      "SQL_HENO", "XLS_HENO", "Δ_HENO",
                                      "SQL_RNO",  "XLS_RNO",  "Δ_RNO",
                                      "SQL_RNF",  "XLS_RNF",  "Δ_RNF",
                                      "SQL_HENF", "XLS_HENF", "Δ_HENF",
                                      "SQL_FEST", "XLS_FEST", "Δ_FEST",
                                      "SQL_TOTAL","XLS_TOTAL","Δ_TOTAL"]
                        _view_cols = [c for c in _view_cols if c in _recon_df.columns]
                        st.caption(
                            "SQL_DO = SALBAS nómina (horas contratadas). "
                            "XLS_DO+DOM = DO+DOM del archivo (horas marcadas en reloj). "
                            "Las columnas Δ muestran SQL − Excel; valores negativos = SQL da menos."
                        )
                        st.dataframe(
                            _recon_df[_view_cols].style.format(
                                {c: "{:.2f}" for c in _view_cols if c != "Mes"}
                            ).applymap(
                                lambda v: "color: #d62728" if isinstance(v, float) and abs(v) > 50 else "",
                                subset=[c for c in _view_cols if c.startswith("Δ_")]
                            ),
                            use_container_width=True,
                            hide_index=True,
                        )

                    if len(_sql_mes_opts) < 2:
                        st.warning("Necesitas al menos 2 meses con datos para comparar escenarios.")
                    else:
                        st.markdown("#### Comparar escenarios por mes completo")
                        _e1, _e2 = st.columns(2)
                        with _e1:
                            _mes_a = st.selectbox("Escenario A (Mes)", _sql_mes_opts, index=max(0, len(_sql_mes_opts) - 2), key="sql_cmp_mes_a")
                        with _e2:
                            _mes_b = st.selectbox("Escenario B (Mes)", _sql_mes_opts, index=len(_sql_mes_opts) - 1, key="sql_cmp_mes_b")

                        _esc_a_df = sql_hist_df[sql_hist_df["Mes"] == _mes_a].copy()
                        _esc_b_df = sql_hist_df[sql_hist_df["Mes"] == _mes_b].copy()

                        # Alineación estricta por mes: solo documentos presentes en ese mes del archivo.
                        if "Documento" in dff.columns and "Mes" in dff.columns and "Documento" in _esc_a_df.columns:
                            _dff_docs_norm = dff.copy()
                            _dff_docs_norm["Documento"] = _normalize_doc_series(_dff_docs_norm["Documento"])
                            _docs_file_a = set(_dff_docs_norm[_dff_docs_norm["Mes"] == _mes_a]["Documento"].tolist())
                            _docs_file_b = set(_dff_docs_norm[_dff_docs_norm["Mes"] == _mes_b]["Documento"].tolist())
                            _esc_a_df = _esc_a_df[_esc_a_df["Documento"].isin(_docs_file_a)].copy()
                            _esc_b_df = _esc_b_df[_esc_b_df["Documento"].isin(_docs_file_b)].copy()

                        st.markdown("##### Conteo de filas/documentos por escenario")
                        _cnt = pd.DataFrame(
                            [
                                {
                                    "Escenario": f"A ({_mes_a})",
                                    "Filas": len(_esc_a_df),
                                    "Documentos": _esc_a_df["Documento"].nunique() if "Documento" in _esc_a_df.columns else 0,
                                },
                                {
                                    "Escenario": f"B ({_mes_b})",
                                    "Filas": len(_esc_b_df),
                                    "Documentos": _esc_b_df["Documento"].nunique() if "Documento" in _esc_b_df.columns else 0,
                                },
                            ]
                        )
                        st.dataframe(_cnt, use_container_width=True, hide_index=True)

                        _esc_a_totals = _esc_a_df.reindex(columns=_cmp_cols, fill_value=0).sum()
                        _esc_b_totals = _esc_b_df.reindex(columns=_cmp_cols, fill_value=0).sum()

                        _cmp_esc = pd.DataFrame({
                            "Tipo": _cmp_cols,
                            f"Escenario A ({_mes_a})": [_esc_a_totals[c] for c in _cmp_cols],
                            f"Escenario B ({_mes_b})": [_esc_b_totals[c] for c in _cmp_cols],
                        })
                        _cmp_esc["Diferencia (A-B)"] = (_cmp_esc[f"Escenario A ({_mes_a})"] - _cmp_esc[f"Escenario B ({_mes_b})"]).round(2)
                        _cmp_esc["Variación %"] = _cmp_esc.apply(
                            lambda r: 0 if r[f"Escenario B ({_mes_b})"] == 0 else round((r["Diferencia (A-B)"] / r[f"Escenario B ({_mes_b})"]) * 100, 2),
                            axis=1,
                        )

                        st.dataframe(
                            _cmp_esc.style.format({c: "{:.2f}" for c in _cmp_esc.columns if c != "Tipo"}),
                            use_container_width=True,
                            hide_index=True,
                        )

                        st.markdown("##### Auditoría por documento (TOTAL)")
                        _doc_a = _esc_a_df.groupby("Documento", as_index=False)["TOTAL"].sum().rename(columns={"TOTAL": f"TOTAL_A_{_mes_a}"})
                        _doc_b = _esc_b_df.groupby("Documento", as_index=False)["TOTAL"].sum().rename(columns={"TOTAL": f"TOTAL_B_{_mes_b}"})
                        _doc_cmp = _doc_a.merge(_doc_b, on="Documento", how="outer").fillna(0)
                        _doc_cmp["Dif (A-B)"] = (_doc_cmp[f"TOTAL_A_{_mes_a}"] - _doc_cmp[f"TOTAL_B_{_mes_b}"]).round(2)
                        _doc_cmp = _doc_cmp.sort_values("Dif (A-B)", ascending=False)
                        st.dataframe(_doc_cmp, use_container_width=True, hide_index=True)

                        _total_row = _cmp_esc[_cmp_esc["Tipo"] == "TOTAL"].iloc[0]
                        _total_diff = _total_row["Diferencia (A-B)"]
                        _total_pct = _total_row["Variación %"]
                        if _total_diff > 0:
                            st.success(f"TOTAL: {_mes_a} subió {_total_diff:,.2f} horas frente a {_mes_b} ({_total_pct:.2f}%).")
                        elif _total_diff < 0:
                            st.warning(f"TOTAL: {_mes_a} bajó {abs(_total_diff):,.2f} horas frente a {_mes_b} ({_total_pct:.2f}%).")
                        else:
                            st.info(f"TOTAL: {_mes_a} y {_mes_b} no presentan variación en horas.")

                        st.caption(
                            "Nota: la diferencia de TOTAL entre SQL y Excel se origina principalmente en DO. "
                            "SQL DO (SALBAS) = horas contratadas; Excel DO+DOM = horas marcadas en reloj. "
                            "La diferencia residual ≈ ausencias pagadas (incapacidades, compensatorios, etc.)."
                        )

# ── Tab: Por Tienda ──
with tabs[6]:
    st.subheader("🏪 Análisis por Tienda / Grupo")
    st.caption("Agrupación basada en el campo 'Grupo' del documento Siesa Access.")

    _hora_cols_tienda = [c for c in ["DO", "RNO", "HEDO", "HENO", "DOM",
                                      "RNF", "HEDF", "HENF", "FEST", "RNDOM",
                                      "RDOM", "ODOM", "ORNF", "OEDF", "OFEST", "TOTAL"]
                         if c in dff.columns]

    t_tienda = (
        dff.groupby("Grupo")[_hora_cols_tienda]
        .sum()
        .reset_index()
        .rename(columns={"Grupo": "Tienda"})
        .sort_values("TOTAL" if "TOTAL" in _hora_cols_tienda else _hora_cols_tienda[-1],
                     ascending=False)
        .reset_index(drop=True)
    )
    _total_col_t = "TOTAL" if "TOTAL" in t_tienda.columns else (_hora_cols_tienda[-1] if _hora_cols_tienda else None)

    # ── KPIs de tiendas ──
    if not t_tienda.empty and _total_col_t:
        top_t = t_tienda.iloc[0]
        _prom = t_tienda[_total_col_t].mean()
        _ti1, _ti2, _ti3 = st.columns(3)
        _ti1.metric("🥇 Tienda con más horas", str(top_t["Tienda"]), f"{top_t[_total_col_t]:,.2f} h")
        _ti2.metric("Número de tiendas", len(t_tienda))
        _ti3.metric("Promedio horas / tienda", f"{_prom:,.2f}")

    # ── Gráfica ranking horizontal ──
    if not t_tienda.empty and _total_col_t:
        _fig_tienda_h = max(350, len(t_tienda) * 42 + 100)
        fig_rank = px.bar(
            t_tienda.sort_values(_total_col_t, ascending=True),
            x=_total_col_t,
            y="Tienda",
            orientation="h",
            color=_total_col_t,
            color_continuous_scale=[[0, "#F0DDD6"], [0.5, "#C9765F"], [1, "#8A2F1F"]],
            labels={_total_col_t: "Total horas", "Tienda": "Tienda"},
            title="Ranking de tiendas por horas totales",
            height=_fig_tienda_h,
            text=_total_col_t,
        )
        fig_rank.update_traces(texttemplate="%{text:,.2f}", textposition="outside")
        fig_rank.update_layout(
            plot_bgcolor="rgba(0,0,0,0)",
            paper_bgcolor="rgba(0,0,0,0)",
            coloraxis_showscale=False,
            font_size=12,
            margin=dict(l=10, r=60, t=50, b=30),
        )
        st.plotly_chart(fig_rank, use_container_width=True)

    # ── Gráfica barras apiladas: composición de horas por tienda ──
    if not t_tienda.empty and len(_hora_cols_tienda) > 1:
        _comp_cols = [c for c in ["DO", "HEDO", "HENO", "DOM", "HEDF", "HENF", "FEST", "RNO", "RNF"]
                      if c in t_tienda.columns]
        if _comp_cols:
            _melt_t = t_tienda[["Tienda"] + _comp_cols].melt(
                id_vars="Tienda", var_name="Tipo", value_name="Horas"
            )
            _melt_t = _melt_t[_melt_t["Horas"] > 0]
            _melt_t["Tipo_desc"] = _melt_t["Tipo"].map(lambda x: HORA_COLS.get(x, x))
            if not _melt_t.empty:
                fig_comp_t = px.bar(
                    _melt_t,
                    x="Tienda",
                    y="Horas",
                    color="Tipo_desc",
                    barmode="stack",
                    height=420,
                    labels={"Tienda": "Tienda", "Horas": "Horas", "Tipo_desc": "Tipo"},
                    title="Composición de horas por tienda",
                )
                fig_comp_t.update_layout(
                    xaxis_tickangle=-30,
                    plot_bgcolor="rgba(0,0,0,0)",
                    paper_bgcolor="rgba(0,0,0,0)",
                    legend_title="Tipo de hora",
                    font_size=12,
                )
                st.plotly_chart(fig_comp_t, use_container_width=True)

    # ── Tabla resumen por tienda ──
    st.markdown("#### Tabla resumen por tienda")
    t_tienda["#"] = range(1, len(t_tienda) + 1)
    _cols_disp_t = ["#", "Tienda"] + [c for c in t_tienda.columns if c not in ("#", "Tienda")]
    t_tienda_disp = t_tienda[_cols_disp_t]
    _fmt_tienda = {c: "{:.2f}" for c in t_tienda_disp.columns if t_tienda_disp[c].dtype.kind in ("f", "i") and c != "#"}
    st.dataframe(
        t_tienda_disp.style.format(_fmt_tienda),
        use_container_width=True,
        height=min(80 + len(t_tienda_disp) * 36, 500),
    )

    # ── Detalle empleados por tienda ──
    st.divider()
    st.markdown("#### Detalle de empleados por tienda")
    _tiendas_list = ["Todas"] + sorted(dff["Grupo"].dropna().unique().tolist())
    sel_tienda_det = st.selectbox("Seleccionar tienda", _tiendas_list, key="sel_tienda_det")

    _det_src = dff if sel_tienda_det == "Todas" else dff[dff["Grupo"] == sel_tienda_det]
    _emp_tienda = (
        _det_src.groupby(["Nombre", "Grupo"])[_hora_cols_tienda]
        .sum()
        .reset_index()
        .rename(columns={"Grupo": "Tienda"})
        .sort_values(_total_col_t if _total_col_t else "Nombre", ascending=False)
        .reset_index(drop=True)
    )
    _fmt_emp_t = {c: "{:.2f}" for c in _emp_tienda.columns if _emp_tienda[c].dtype.kind in ("f", "i")}
    st.dataframe(
        _emp_tienda.style.format(_fmt_emp_t),
        use_container_width=True,
        height=min(80 + len(_emp_tienda) * 36, 500),
    )
    _xlsx_emp_tienda = df_to_excel_grouped(_emp_tienda, "Empleados por tienda", group_col="Tienda")
    st.download_button(
        "⬇ Descargar Excel (empleados por tienda)",
        data=_xlsx_emp_tienda,
        file_name=f"empleados_{sel_tienda_det.replace(' ','_')}.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        key="dl_emp_tienda",
    )

    # ── Descarga Excel ──
    _t_tienda_exp = t_tienda_disp.drop(columns=["#"])
    _xlsx_tienda = df_to_excel_grouped(_t_tienda_exp, "Por tienda", group_col="Tienda")
    st.download_button(
        "⬇ Descargar Excel (resumen por tienda)",
        data=_xlsx_tienda,
        file_name="reporte_por_tienda.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        key="dl_tienda",
    )

# ─────────────────────────────────────────────
st.caption("Desarrollado con Streamlit · Datos: Siesa Access")
