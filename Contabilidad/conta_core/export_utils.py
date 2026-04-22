import io

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

COLS_HORAS_EXTRA = ["HEDO", "HENO", "HEDF", "HENF"]
LIMITE_DIARIO_HE = 2.0
LIMITE_SEMANAL_HE = 12.0


def calcular_cumplimiento(df: pd.DataFrame) -> tuple[pd.DataFrame, pd.DataFrame]:
    extras_cols = [c for c in COLS_HORAS_EXTRA if c in df.columns]

    work = df.copy()
    work["_HE_Total"] = work[extras_cols].sum(axis=1) if extras_cols else 0.0

    grp_dia = (
        work.groupby(["Nombre", "Fecha"])["_HE_Total"]
        .sum()
        .reset_index()
        .rename(columns={"_HE_Total": "H. Extra del dia"})
    )
    grp_dia["Fecha"] = grp_dia["Fecha"].dt.strftime("%d/%m/%Y")
    grp_dia["Limite diario"] = LIMITE_DIARIO_HE
    grp_dia["Exceso diario"] = (grp_dia["H. Extra del dia"] - LIMITE_DIARIO_HE).clip(lower=0).round(2)
    grp_dia["Estado"] = grp_dia["H. Extra del dia"].apply(
        lambda x: "EXCEDE LIMITE" if x > LIMITE_DIARIO_HE else "OK"
    )
    grp_dia = grp_dia[grp_dia["H. Extra del dia"] > 0].reset_index(drop=True)

    if "Semana_etiqueta" in work.columns:
        grp_sem = (
            work.groupby(["Nombre", "Semana_etiqueta"])["_HE_Total"]
            .sum()
            .reset_index()
            .rename(columns={"Semana_etiqueta": "Semana", "_HE_Total": "H. Extra semana"})
        )
    else:
        grp_sem = pd.DataFrame(columns=["Nombre", "Semana", "H. Extra semana"])

    grp_sem["Limite semanal"] = LIMITE_SEMANAL_HE
    grp_sem["Exceso semanal"] = (grp_sem["H. Extra semana"] - LIMITE_SEMANAL_HE).clip(lower=0).round(2)
    grp_sem["Estado"] = grp_sem["H. Extra semana"].apply(
        lambda x: "EXCEDE LIMITE" if x > LIMITE_SEMANAL_HE else "OK"
    )
    grp_sem = grp_sem[grp_sem["H. Extra semana"] > 0].reset_index(drop=True)

    return grp_dia, grp_sem


def excel_cumplimiento(df_diario: pd.DataFrame, df_semanal: pd.DataFrame) -> bytes:
    wb = Workbook()

    color_ok = "E8F5E9"
    color_excede = "FFEBEE"
    color_header = "8A2F1F"
    thin = Side(style="thin", color="CCCCCC")
    brd = Border(left=thin, right=thin, top=thin, bottom=thin)

    def _write_sheet(ws, df, titulo):
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(df.columns))
        title_cell = ws.cell(row=1, column=1, value=titulo)
        title_cell.font = Font(bold=True, size=12, color="FFFFFF")
        title_cell.fill = PatternFill("solid", fgColor=color_header)
        title_cell.alignment = Alignment(horizontal="center")

        hdr_font = Font(bold=True, color="FFFFFF", size=10)
        hdr_fill = PatternFill("solid", fgColor="5D2010")
        for ci, col in enumerate(df.columns, 1):
            c = ws.cell(row=2, column=ci, value=col)
            c.font = hdr_font
            c.fill = hdr_fill
            c.alignment = Alignment(horizontal="center")
            c.border = brd
        ws.freeze_panes = "A3"

        for ri, (_, row) in enumerate(df.iterrows(), 3):
            es_exceso = "EXCEDE" in str(row.get("Estado", ""))
            fill = PatternFill("solid", fgColor=color_excede if es_exceso else color_ok)
            for ci, val in enumerate(row, 1):
                c = ws.cell(row=ri, column=ci, value=val)
                c.fill = fill
                c.border = brd
                c.font = Font(size=9, bold=es_exceso, color="B71C1C" if es_exceso else "1B5E20")
                if isinstance(val, float):
                    c.number_format = "0.00"

        for ci, col in enumerate(df.columns, 1):
            max_len = max(len(str(col)), *(len(str(v)) for v in df.iloc[:, ci - 1]))
            ws.column_dimensions[get_column_letter(ci)].width = min(max_len + 2, 35)

    ws1 = wb.active
    ws1.title = "Incumplimiento Diario"
    _write_sheet(ws1, df_diario, "Reporte de cumplimiento - Horas extras diarias (max. 2 h/dia)")

    ws2 = wb.create_sheet("Incumplimiento Semanal")
    _write_sheet(ws2, df_semanal, "Reporte de cumplimiento - Horas extras semanales (max. 12 h/semana)")

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()


def df_to_excel_grouped(df: pd.DataFrame, title: str, group_col: str = "Nombre") -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = title[:30]

    color_header = "8A2F1F"
    color_group = "F0DDD6"
    color_total = "FFF3F0"

    thin = Side(style="thin", color="CCCCCC")
    brd = Border(left=thin, right=thin, top=thin, bottom=thin)

    cols = list(df.columns)
    has_group = group_col in cols
    display_cols = [c for c in cols if c != group_col] if has_group else cols

    hdr_fill = PatternFill("solid", fgColor=color_header)
    hdr_font = Font(bold=True, color="FFFFFF", size=10)
    for ci, col in enumerate(display_cols, 1):
        cell = ws.cell(row=1, column=ci, value=col)
        cell.fill = hdr_fill
        cell.font = hdr_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = brd
    ws.row_dimensions[1].height = 18
    ws.freeze_panes = "A2"

    grp_fill = PatternFill("solid", fgColor=color_group)
    grp_font = Font(bold=True, color="3D1A0E", size=10)
    data_font = Font(size=9)
    tot_fill = PatternFill("solid", fgColor=color_total)
    tot_font = Font(bold=True, size=9)

    row_num = 2
    groups = df.groupby(group_col, sort=False) if has_group else [(None, df)]
    for emp, grp in groups:
        if has_group:
            ws.merge_cells(start_row=row_num, start_column=1, end_row=row_num, end_column=len(display_cols))
            cell = ws.cell(row=row_num, column=1, value=str(emp))
            cell.fill = grp_fill
            cell.font = grp_font
            cell.alignment = Alignment(horizontal="left", vertical="center")
            ws.row_dimensions[row_num].height = 16
            row_num += 1

        num_cols_names = [c for c in display_cols if grp[c].dtype.kind in ("f", "i")]

        for _, data_row in grp.iterrows():
            for ci, col in enumerate(display_cols, 1):
                val = data_row[col]
                cell = ws.cell(row=row_num, column=ci, value=val)
                cell.font = data_font
                cell.border = brd
                if isinstance(val, float):
                    cell.number_format = "0.00"
            row_num += 1

        if has_group:
            for ci, col in enumerate(display_cols, 1):
                if col in num_cols_names:
                    val = grp[col].sum()
                    cell = ws.cell(row=row_num, column=ci, value=val)
                    cell.number_format = "0.00"
                else:
                    cell = ws.cell(row=row_num, column=ci, value="TOTAL" if ci == 1 else "")
                cell.fill = tot_fill
                cell.font = tot_font
                cell.border = brd
            row_num += 2

    for ci, col in enumerate(display_cols, 1):
        max_len = len(str(col))
        for row_cells in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=ci, max_col=ci):
            for cell in row_cells:
                if cell.value:
                    max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[get_column_letter(ci)].width = min(max_len + 2, 32)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()
