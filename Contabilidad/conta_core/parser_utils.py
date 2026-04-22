import openpyxl
import pandas as pd

HORA_COLS = {
    "JORNADA": "Jornada ordinaria",
    "DO": "D. ordinario",
    "RNO": "Recargo nocturno ord.",
    "HEDO": "H. extra diurna ord.",
    "HENO": "H. extra nocturna ord.",
    "DOM": "Dominical",
    "RNF": "Rec. noct. festivo",
    "HEDF": "H. extra diurna festivo",
    "HENF": "H. extra noct. festivo",
    "FEST": "Festivo",
    "RNDOM": "Rec. noct. dominical",
    "RDOM": "Rec. dominical",
    "ODOM": "Hora extra dominical",
    "ORNF": "Otra rec. noct. festivo",
    "OEDF": "Otra h. ext. diurna fest.",
    "OFEST": "Otras horas festivo",
    "TOTAL": "Total horas",
}

DIAS = {
    "lunes",
    "martes",
    "miércoles",
    "miercoles",
    "jueves",
    "viernes",
    "sábado",
    "sabado",
    "domingo",
}


def parse_excel_file(path_or_file) -> pd.DataFrame:
    if hasattr(path_or_file, "read"):
        wb = openpyxl.load_workbook(path_or_file, data_only=True)
    else:
        wb = openpyxl.load_workbook(str(path_or_file), data_only=True)

    ws = wb["ReporteXML"]
    rows = list(ws.iter_rows(values_only=True))

    records = []
    current_nombre = None
    current_documento = None
    current_grupo = None
    col_map = {}

    for row in rows:
        vals = {i: v for i, v in enumerate(row) if v is not None}
        if not vals:
            continue

        for value in vals.values():
            sval = str(value).strip()
            if sval.startswith("Nombre:"):
                current_nombre = sval.replace("Nombre:", "").strip()
                col_map = {}
                break

        for value in vals.values():
            sval = str(value).strip()
            if sval.startswith("Documento:"):
                current_documento = sval.replace("Documento:", "").strip()
                break

        for value in vals.values():
            sval = str(value).strip()
            if sval.startswith("Grupo:"):
                current_grupo = " ".join(sval.replace("Grupo:", "").split())
                break

        row_str_vals = [str(v).strip() for v in vals.values()]
        if "Fecha" in row_str_vals and "TOTAL" in row_str_vals:
            col_map = {}
            for idx, value in vals.items():
                if value is not None:
                    col_map[idx] = str(value).strip().replace("\n", " ")
            continue

        if col_map and current_nombre:
            first_val = None
            first_idx = None
            for idx in sorted(vals.keys()):
                if vals[idx] is not None:
                    first_val = str(vals[idx]).strip().lower()
                    first_idx = idx
                    break

            if first_val and first_val in DIAS:
                record = {
                    "Nombre": current_nombre,
                    "Documento": current_documento,
                    "Grupo": current_grupo,
                }
                for col_idx, col_name in col_map.items():
                    raw = vals.get(col_idx)
                    record[col_name] = str(raw).strip() if raw is not None and str(raw).strip() not in ("", " ") else None

                record["Día"] = str(vals.get(first_idx, "")).strip().capitalize()
                records.append(record)

    if not records:
        return pd.DataFrame()

    df = pd.DataFrame(records)

    if "Fecha" in df.columns:
        raw_fecha = df["Fecha"].astype(str).str.strip()
        df["Fecha"] = pd.to_datetime(raw_fecha, errors="coerce", format="%m/%d/%Y")
        mask = df["Fecha"].isna()
        if mask.any():
            df.loc[mask, "Fecha"] = pd.to_datetime(raw_fecha[mask], errors="coerce")

    if "Fecha" in df.columns:
        df["Semana"] = df["Fecha"].dt.isocalendar().week.astype("Int64")
        df["Semana_etiqueta"] = df["Fecha"].dt.to_period("W").astype(str)
        df["Mes"] = df["Fecha"].dt.strftime("%B %Y")
        df["Mes_num"] = df["Fecha"].dt.month

    for col in list(HORA_COLS.keys()) + ["TOTAL"]:
        if col in df.columns:
            df[col] = pd.to_numeric(df[col], errors="coerce").fillna(0.0)

    total_col = next((c for c in df.columns if str(c).upper() == "TOTAL"), None)
    if total_col and total_col != "TOTAL":
        df["TOTAL"] = df[total_col]

    if "Fecha" in df.columns:
        df = df[df["Fecha"].notna()].copy()

    return df


def prepare_loaded_dataframe(df: pd.DataFrame) -> tuple[pd.DataFrame, list[str]]:
    prepared = df.copy()
    if "Grupo" in prepared.columns:
        prepared["Grupo"] = prepared["Grupo"].astype(str).str.strip().str.replace(r"\s+", " ", regex=True)
        prepared["Grupo"] = prepared["Grupo"].replace("None", "(Sin grupo)").fillna("(Sin grupo)")

    horas_disponibles = [c for c in HORA_COLS if c in prepared.columns]
    if "TOTAL" in prepared.columns and "TOTAL" not in horas_disponibles:
        horas_disponibles.append("TOTAL")

    return prepared, horas_disponibles
