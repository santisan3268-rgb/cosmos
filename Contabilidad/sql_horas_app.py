import os
from datetime import date

import pandas as pd
import pytds
import streamlit as st
import openpyxl

HORA_COLS = [
    "DO",
    "RNO",
    "HEDO",
    "HENO",
    "DOM",
    "RNF",
    "HEDF",
    "HENF",
    "FEST",
    "RNDOM",
    "RDOM",
    "ODOM",
    "ORNF",
    "OEDF",
    "OFEST",
    "TOTAL",
]


def _norm_text(value: str) -> str:
    return " ".join(str(value or "").upper().split())


def _to_bool(value, default: bool = True) -> bool:
    if value is None:
        return default
    if isinstance(value, bool):
        return value
    return str(value).strip().lower() in {"1", "true", "yes", "y", "si", "on"}


@st.cache_data(show_spinner="Leyendo Excel de Siesa...")
def parse_excel_siesa(path_or_file) -> pd.DataFrame:
    if hasattr(path_or_file, "read"):
        wb = openpyxl.load_workbook(path_or_file, data_only=True)
    else:
        wb = openpyxl.load_workbook(str(path_or_file), data_only=True)

    ws = wb["ReporteXML"]
    rows = list(ws.iter_rows(values_only=True))

    records = []
    current_nombre = None
    current_documento = None
    current_grupo = None
    col_map = {}

    dias = {"lunes", "martes", "miercoles", "miércoles", "jueves", "viernes", "sabado", "sábado", "domingo"}

    for row in rows:
        vals = {i: v for i, v in enumerate(row) if v is not None}
        if not vals:
            continue

        for v in vals.values():
            sv = str(v).strip()
            if sv.startswith("Nombre:"):
                current_nombre = sv.replace("Nombre:", "").strip()
                col_map = {}
                break

        for v in vals.values():
            sv = str(v).strip()
            if sv.startswith("Documento:"):
                current_documento = sv.replace("Documento:", "").strip()
                break

        for v in vals.values():
            sv = str(v).strip()
            if sv.startswith("Grupo:"):
                current_grupo = " ".join(sv.replace("Grupo:", "").split())
                break

        row_str_vals = [str(v).strip() for v in vals.values()]
        if "Fecha" in row_str_vals and "TOTAL" in row_str_vals:
            col_map = {idx: str(v).strip().replace("\n", " ") for idx, v in vals.items()}
            continue

        if col_map and current_nombre:
            first_idx = min(vals.keys())
            first_val = str(vals[first_idx]).strip().lower()
            if first_val in dias:
                rec = {
                    "Nombre": current_nombre,
                    "Documento": current_documento,
                    "Grupo": current_grupo,
                }
                for col_idx, col_name in col_map.items():
                    raw = vals.get(col_idx)
                    rec[col_name] = str(raw).strip() if raw is not None and str(raw).strip() else None
                records.append(rec)

    if not records:
        return pd.DataFrame()

    df = pd.DataFrame(records)
    if "Fecha" in df.columns:
        df["Fecha"] = pd.to_datetime(df["Fecha"], errors="coerce")

    for col in HORA_COLS:
        if col in df.columns:
            df[col] = pd.to_numeric(df[col], errors="coerce").fillna(0.0)

    df = df[df["Fecha"].notna()].copy()
    df["Mes"] = df["Fecha"].dt.to_period("M").astype(str)
    return df


def connect_sql(server: str, port: int, db_name: str, user: str, password: str):
    validate_host = _to_bool(_get_secret_or_env("DB_VALIDATE_HOST", True), True)
    return pytds.connect(
        server=server,
        port=port,
        database=db_name,
        user=user,
        password=password,
        validate_host=validate_host,
    )


def _get_secret_or_env(key: str, default=None):
    try:
        if key in st.secrets:
            return st.secrets[key]
    except Exception:
        pass
    return os.getenv(key, default)


def get_db_config() -> dict:
    return {
        "server": _get_secret_or_env("DB_SERVER"),
        "port": int(_get_secret_or_env("DB_PORT", "1433")),
        "name": _get_secret_or_env("DB_NAME"),
        "user": _get_secret_or_env("DB_USER"),
        "password": _get_secret_or_env("DB_PASSWORD"),
    }


def validate_db_config(cfg: dict) -> list:
    missing = []
    if not cfg.get("server"):
        missing.append("DB_SERVER")
    if not cfg.get("name"):
        missing.append("DB_NAME")
    if not cfg.get("user"):
        missing.append("DB_USER")
    if not cfg.get("password"):
        missing.append("DB_PASSWORD")
    return missing


@st.cache_data(show_spinner=False)
def test_sql_connection(server: str, port: int, db_name: str, user: str, password: str) -> bool:
    conn = connect_sql(server, port, db_name, user, password)
    cur = conn.cursor()
    cur.execute("SELECT 1")
    ok = cur.fetchone()[0] == 1
    conn.close()
    return ok


@st.cache_data(show_spinner="Descubriendo tablas candidatas en SQL...")
def discover_sources(server: str, port: int, db_name: str, user: str, password: str) -> pd.DataFrame:
    conn = connect_sql(server, port, db_name, user, password)
    cur = conn.cursor()
    query = """
    SELECT s.name AS schema_name, t.name AS table_name, c.name AS column_name
    FROM sys.tables t
    JOIN sys.schemas s ON s.schema_id = t.schema_id
    JOIN sys.columns c ON c.object_id = t.object_id
    WHERE LOWER(t.name) LIKE '%nom%'
       OR LOWER(t.name) LIKE '%hora%'
       OR LOWER(t.name) LIKE '%turn%'
       OR LOWER(t.name) LIKE '%emple%'
       OR LOWER(t.name) LIKE '%labor%'
    """
    cur.execute(query)
    rows = cur.fetchall()
    conn.close()

    df = pd.DataFrame(rows, columns=["schema_name", "table_name", "column_name"])

    if df.empty:
        return df

    score = []
    for _, row in df.iterrows():
        table = row["table_name"].lower()
        col = row["column_name"].lower()
        s = 0
        if table == "w0602_movto_nomina":
            s += 10
        if table in {"w0501_conceptos", "t200_mm_terceros", "w0600_docto_nomina", "w0601_docto_nomina_emp"}:
            s += 6
        for token in ["hora", "fecha", "concept", "tercero", "period", "docto"]:
            if token in col:
                s += 1
        score.append(s)

    df["score"] = score
    ranking = (
        df.groupby(["schema_name", "table_name"], as_index=False)
        .agg(score=("score", "sum"), cols_match=("column_name", "count"))
        .sort_values(["score", "cols_match"], ascending=False)
        .reset_index(drop=True)
    )
    return ranking


def classify_concept(concept_id: str, concept_desc: str, concept_abbr: str) -> str:
    text = _norm_text(f"{concept_id} {concept_desc} {concept_abbr}")

    if "RECARGO NOCTURNO" in text and "DOM" not in text and "FEST" not in text:
        return "RNO"
    if "HORA EXTRA DIURNA" in text and "DOM" not in text and "FEST" not in text:
        return "HEDO"
    if "HORA EXTRA NOCTURNA" in text and "DOM" not in text and "FEST" not in text:
        return "HENO"
    if "DOMING" in text or "DOMINICAL" in text:
        return "DOM"
    if "FEST" in text:
        return "FEST"
    if "SALARIO BASICO" in text or "JORNADA" in text:
        return "DO"
    return "OTRAS"


@st.cache_data(show_spinner="Consultando horas históricas en SQL...")
def fetch_sql_hours(server: str, port: int, db_name: str, user: str, password: str, fecha_ini: date, fecha_fin: date) -> pd.DataFrame:
    conn = connect_sql(server, port, db_name, user, password)
    cur = conn.cursor()
    query = """
    SELECT
        CAST(m.c0602_fecha AS date) AS Fecha,
        t.f200_nit AS Documento,
        LTRIM(RTRIM(CONCAT(ISNULL(t.f200_nombres, ''), ' ', ISNULL(t.f200_apellido1, ''), ' ', ISNULL(t.f200_apellido2, '')))) AS Nombre,
        c.c0501_id AS ConceptoID,
        c.c0501_descripcion AS Concepto,
        c.c0501_abreviatura AS Abreviatura,
        SUM(ISNULL(m.c0602_horas, 0)) AS Horas
    FROM dbo.w0602_movto_nomina m
    INNER JOIN dbo.w0501_conceptos c ON c.c0501_rowid = m.c0602_rowid_concepto
    LEFT JOIN dbo.t200_mm_terceros t ON t.f200_rowid = m.c0602_rowid_tercero AND t.f200_id_cia = m.c0602_id_cia
    WHERE m.c0602_horas IS NOT NULL
      AND m.c0602_horas > 0
      AND CAST(m.c0602_fecha AS date) BETWEEN %s AND %s
    GROUP BY
        CAST(m.c0602_fecha AS date),
        t.f200_nit,
        t.f200_nombres,
        t.f200_apellido1,
        t.f200_apellido2,
        c.c0501_id,
        c.c0501_descripcion,
        c.c0501_abreviatura
    """
    cur.execute(query, (fecha_ini, fecha_fin))
    rows = cur.fetchall()
    conn.close()

    raw = pd.DataFrame(
        rows,
        columns=["Fecha", "Documento", "Nombre", "ConceptoID", "Concepto", "Abreviatura", "Horas"],
    )

    if raw.empty:
        return raw

    raw["Fecha"] = pd.to_datetime(raw["Fecha"])
    raw["Categoria"] = raw.apply(
        lambda r: classify_concept(r["ConceptoID"], r["Concepto"], r["Abreviatura"]),
        axis=1,
    )

    grouped = (
        raw.groupby(["Fecha", "Documento", "Nombre", "Categoria"], as_index=False)["Horas"]
        .sum()
        .pivot_table(index=["Fecha", "Documento", "Nombre"], columns="Categoria", values="Horas", fill_value=0)
        .reset_index()
    )

    grouped.columns = [str(c) for c in grouped.columns]
    for col in ["DO", "RNO", "HEDO", "HENO", "DOM", "FEST", "OTRAS"]:
        if col not in grouped.columns:
            grouped[col] = 0.0

    grouped["TOTAL"] = grouped[["DO", "RNO", "HEDO", "HENO", "DOM", "FEST", "OTRAS"]].sum(axis=1)
    grouped["Mes"] = grouped["Fecha"].dt.to_period("M").astype(str)
    return grouped


def compare_months(df_a: pd.DataFrame, label_a: str, df_b: pd.DataFrame, label_b: str) -> pd.DataFrame:
    a = df_a.groupby("Mes", as_index=False)["TOTAL"].sum().rename(columns={"TOTAL": label_a})
    b = df_b.groupby("Mes", as_index=False)["TOTAL"].sum().rename(columns={"TOTAL": label_b})
    out = a.merge(b, on="Mes", how="outer").fillna(0)
    out["Diferencia"] = out[label_a] - out[label_b]
    out["Variacion_%"] = out.apply(
        lambda r: 0.0 if r[label_b] == 0 else (r[label_a] - r[label_b]) / r[label_b] * 100,
        axis=1,
    )
    return out.sort_values("Mes").reset_index(drop=True)


st.set_page_config(page_title="Contabilidad SQL Horas", layout="wide")
st.title("Contabilidad - Validacion SQL de Horas Laboradas")
st.caption("Fuente principal propuesta: dbo.w0602_movto_nomina + dbo.w0501_conceptos + dbo.t200_mm_terceros")

db_cfg = get_db_config()
missing_cfg = validate_db_config(db_cfg)

if missing_cfg:
    st.error(
        "Faltan variables de entorno/secrets para conexion SQL: "
        + ", ".join(missing_cfg)
        + "."
    )
    st.info("Configuralas en variables de entorno o en .streamlit/secrets.toml. La conexion es automatica y no se expone en pantalla.")
    st.stop()

try:
    sql_ok = test_sql_connection(db_cfg["server"], db_cfg["port"], db_cfg["name"], db_cfg["user"], db_cfg["password"])
except Exception as e:
    st.error(f"No se pudo conectar automaticamente a SQL: {e}")
    st.stop()

with st.sidebar:
    st.subheader("Conexion SQL")
    st.success("Conexion automatica activa" if sql_ok else "Conexion no disponible")
    st.caption(f"Servidor: {db_cfg['server']}")
    st.caption(f"Base: {db_cfg['name']}")
    st.caption(f"Usuario: {db_cfg['user']}")

    st.divider()
    st.subheader("Periodo")
    fecha_ini = st.date_input("Desde", value=date(2024, 1, 1))
    fecha_fin = st.date_input("Hasta", value=date.today())

col_a, col_b = st.columns([1, 1])
with col_a:
    if st.button("Descubrir tablas candidatas", use_container_width=True):
        try:
            ranking = discover_sources(db_cfg["server"], db_cfg["port"], db_cfg["name"], db_cfg["user"], db_cfg["password"])
            st.success("Tablas candidatas detectadas.")
            st.dataframe(ranking.head(20), use_container_width=True, hide_index=True)
        except Exception as e:
            st.error(f"No fue posible consultar metadatos SQL: {e}")

with col_b:
    if st.button("Consultar horas historicas SQL", type="primary", use_container_width=True):
        try:
            sql_df = fetch_sql_hours(db_cfg["server"], db_cfg["port"], db_cfg["name"], db_cfg["user"], db_cfg["password"], fecha_ini, fecha_fin)
            if sql_df.empty:
                st.warning("No se encontraron horas en el rango seleccionado.")
            else:
                st.session_state["sql_df_horas"] = sql_df
                st.success(f"Consulta completada: {len(sql_df):,} filas.")
        except Exception as e:
            st.error(f"Error consultando horas en SQL: {e}")

sql_df = st.session_state.get("sql_df_horas", pd.DataFrame())

st.divider()
st.subheader("1) Comparacion mensual SQL")
if sql_df.empty:
    st.info("Ejecuta primero 'Consultar horas historicas SQL'.")
else:
    sql_month = sql_df.groupby("Mes", as_index=False)[["DO", "RNO", "HEDO", "HENO", "DOM", "FEST", "TOTAL"]].sum()
    st.dataframe(sql_month.sort_values("Mes"), use_container_width=True, hide_index=True)

    m1, m2 = st.columns(2)
    with m1:
        mes_1 = st.selectbox("Mes A", sorted(sql_df["Mes"].unique()), key="mes_sql_a")
    with m2:
        mes_2 = st.selectbox("Mes B", sorted(sql_df["Mes"].unique()), key="mes_sql_b")

    c1 = sql_df[sql_df["Mes"] == mes_1]["TOTAL"].sum()
    c2 = sql_df[sql_df["Mes"] == mes_2]["TOTAL"].sum()
    diff = c1 - c2
    pct = 0.0 if c2 == 0 else (diff / c2) * 100

    k1, k2, k3 = st.columns(3)
    k1.metric(f"TOTAL {mes_1}", f"{c1:,.2f}")
    k2.metric(f"TOTAL {mes_2}", f"{c2:,.2f}")
    k3.metric("Diferencia", f"{diff:,.2f}", delta=f"{pct:,.2f}%")

st.divider()
st.subheader("2) Validacion SQL vs documentos contables")
excel_files = st.file_uploader(
    "Carga uno o varios Excel contables (Marzo 2026, 1-31 Octubre, Febrero FYC)",
    type=["xlsx"],
    accept_multiple_files=True,
)

if excel_files:
    excel_parts = []
    for f in excel_files:
        try:
            part = parse_excel_siesa(f)
            if not part.empty:
                excel_parts.append(part)
        except Exception as e:
            st.warning(f"No se pudo leer {f.name}: {e}")

    if excel_parts:
        excel_df = pd.concat(excel_parts, ignore_index=True)
        st.success(f"Documentos leidos: {len(excel_parts)} archivos, {len(excel_df):,} filas.")

        excel_month = excel_df.groupby("Mes", as_index=False)[[c for c in HORA_COLS if c in excel_df.columns]].sum()
        st.markdown("Resumen mensual desde documentos")
        st.dataframe(excel_month.sort_values("Mes"), use_container_width=True, hide_index=True)

        if not sql_df.empty:
            cmp_df = compare_months(sql_df, "SQL_TOTAL", excel_df, "DOC_TOTAL")
            st.markdown("Comparacion mensual SQL vs documentos")
            st.dataframe(cmp_df, use_container_width=True, hide_index=True)

            abs_diff = cmp_df["Diferencia"].abs().sum()
            if abs_diff < 0.01:
                st.success("Validacion OK: SQL y documentos coinciden en TOTAL mensual dentro del umbral.")
            else:
                st.warning("Se detectaron diferencias entre SQL y documentos. Revisa mapeo de conceptos y filtros.")

            st.caption(
                "Nota tecnica: la extraccion SQL usa dbo.w0602_movto_nomina (horas), "
                "dbo.w0501_conceptos (clasificacion) y dbo.t200_mm_terceros (documento/nombre)."
            )
        else:
            st.info("Para validar contra SQL, primero ejecuta 'Consultar horas historicas SQL'.")
    else:
        st.error("No se pudieron extraer datos de los archivos Excel cargados.")
