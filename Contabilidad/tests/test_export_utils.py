import sys
import unittest
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(ROOT))

from conta_core.export_utils import calcular_cumplimiento, df_to_excel_grouped, excel_cumplimiento


class ExportUtilsTests(unittest.TestCase):
    def test_calcular_cumplimiento_detects_daily_and_weekly_overages(self):
        df = pd.DataFrame(
            {
                "Nombre": ["Ana", "Ana", "Luis"],
                "Fecha": pd.to_datetime(["2026-02-03", "2026-02-04", "2026-02-03"]),
                "Semana_etiqueta": ["2026-02-02/2026-02-08", "2026-02-02/2026-02-08", "2026-02-02/2026-02-08"],
                "HEDO": [3.0, 6.0, 1.0],
                "HENO": [0.0, 5.0, 0.0],
            }
        )

        diario, semanal = calcular_cumplimiento(df)

        self.assertIn("Estado", diario.columns)
        self.assertTrue((diario["Estado"] == "EXCEDE LIMITE").any())
        self.assertTrue((semanal["Estado"] == "EXCEDE LIMITE").any())

    def test_df_to_excel_grouped_generates_workbook_bytes(self):
        df = pd.DataFrame({"Nombre": ["Ana", "Ana"], "DO": [8.0, 7.5], "TOTAL": [8.0, 7.5]})
        content = df_to_excel_grouped(df, "Prueba", group_col="Nombre")
        wb = load_workbook(filename=Path(ROOT / "tests_temp.xlsx") if False else __import__("io").BytesIO(content))
        self.assertEqual(wb.active.title, "Prueba")
        self.assertEqual(wb.active["A1"].value, "DO")

    def test_excel_cumplimiento_creates_two_sheets(self):
        diario = pd.DataFrame({"Nombre": ["Ana"], "Fecha": ["03/02/2026"], "Estado": ["OK"]})
        semanal = pd.DataFrame({"Nombre": ["Ana"], "Semana": ["2026-W06"], "Estado": ["OK"]})
        content = excel_cumplimiento(diario, semanal)
        wb = load_workbook(filename=__import__("io").BytesIO(content))
        self.assertIn("Incumplimiento Diario", wb.sheetnames)
        self.assertIn("Incumplimiento Semanal", wb.sheetnames)


if __name__ == "__main__":
    unittest.main()
